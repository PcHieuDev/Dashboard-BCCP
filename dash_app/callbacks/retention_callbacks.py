# -*- coding: utf-8 -*-
"""
Callbacks xử lý dữ liệu trang Báo cáo Duy trì & Biến động Khách hàng Hiện hữu (/bccp/retention) v2.0.
"""

import sqlite3
import pandas as pd
import dash
from dash import Output, Input, State, html, dcc, dash_table
import dash_bootstrap_components as dbc
from datetime import datetime
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

import sys
from pathlib import Path

# Thêm thư mục gốc vào sys.path để import
project_root = Path(__file__).resolve().parent.parent.parent
if str(project_root) not in sys.path:
    sys.path.insert(0, str(project_root))

from config.settings import DB_PATH
from callbacks.utils import format_revenue
from analytics.retention_metrics import get_khhh_changes_v2, get_weekly_changes

def generate_retention_list_excel(df: pd.DataFrame, title: str, is_roi_bo: bool, period_str: str) -> bytes:
    """Tạo file Excel xuất báo cáo chi tiết biến động khách hàng"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.views.sheetView[0].showGridLines = True
    
    font_name = "Arial"
    
    # Tiêu đề báo cáo
    ws.merge_cells("A1:H1")
    ws["A1"] = f"{title.upper()} - {period_str.upper()}"
    ws["A1"].font = Font(name=font_name, size=14, bold=True, color="1E3A8A")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40
    
    ws["A3"] = f"• Ngày xuất báo cáo: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    ws["A3"].font = Font(name=font_name, size=10, italic=True)
    
    # Headers
    if is_roi_bo:
        headers = ["Tên cụm", "Tên xã", "Mã CMS", "SL gần nhất", "DT gần nhất", "Kỳ gần nhất có DT"]
        cols = ["ten_cum", "ten_bdx", "cms", "sl_gan_nhat", "dt_gan_nhat", "thang_gan_nhat"]
    else:
        headers = ["Tên cụm", "Tên xã", "Mã CMS", "SL kỳ này", "DT kỳ này", "SL kỳ trước", "DT kỳ trước", "Chênh lệch DT"]
        cols = ["ten_cum", "ten_bdx", "cms", "sl_ky_nay", "dt_ky_nay", "sl_ky_truoc", "dt_ky_truoc", "chenh_lech"]
        
    header_row = 5
    ws.row_dimensions[header_row].height = 25
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(name=font_name, size=10, bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    
    for c_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=c_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        
    # Ghi dữ liệu
    data_start = 6
    for r_idx, row in enumerate(df.to_dict('records'), data_start):
        ws.row_dimensions[r_idx].height = 20
        is_odd = (r_idx % 2 == 1)
        row_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid") if is_odd else PatternFill(fill_type=None)
        
        for c_idx, col in enumerate(cols, 1):
            val = row.get(col, None)
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.border = thin_border
            if row_fill.fill_type:
                cell.fill = row_fill
                
            cell.font = Font(name=font_name, size=9)
            if pd.isna(val) or val is None:
                cell.value = "-"
                cell.alignment = Alignment(horizontal="center")
            elif col in ["ten_cum", "ten_bdx"]:
                cell.value = str(val)
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif col == "cms":
                cell.value = str(val)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(name=font_name, size=9, bold=True)
            elif col in ["sl_ky_nay", "sl_ky_truoc", "sl_gan_nhat"]:
                cell.value = int(val)
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col in ["dt_ky_nay", "dt_ky_truoc", "dt_gan_nhat", "chenh_lech"]:
                cell.value = float(val)
                cell.number_format = '#,##0" đ"'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.value = str(val)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
    # Dòng tổng cộng
    total_row = data_start + len(df)
    ws.row_dimensions[total_row].height = 22
    total_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    total_font = Font(name=font_name, size=9, bold=True)
    
    ws.merge_cells(f"A{total_row}:C{total_row}")
    cell_lbl = ws.cell(row=total_row, column=1, value="Tổng cộng")
    cell_lbl.font = Font(name=font_name, size=10, bold=True)
    cell_lbl.alignment = Alignment(horizontal="left", vertical="center")
    
    for c_idx in range(1, len(cols) + 1):
        cell = ws.cell(row=total_row, column=c_idx)
        cell.fill = total_fill
        cell.border = thin_border
        if c_idx > 3:
            cell.font = total_font
            col_letter = get_column_letter(c_idx)
            col_name = cols[c_idx - 1]
            if col_name in ["sl_ky_nay", "sl_ky_truoc", "sl_gan_nhat"]:
                cell.value = f"=SUM({col_letter}{data_start}:{col_letter}{total_row-1})"
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_name in ["dt_ky_nay", "dt_ky_truoc", "dt_gan_nhat", "chenh_lech"]:
                cell.value = f"=SUM({col_letter}{data_start}:{col_letter}{total_row-1})"
                cell.number_format = '#,##0" đ"'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.value = ""
                
    # Điều chỉnh độ rộng cột tự động
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row == 1:
                continue
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    output = io.BytesIO()
    wb.save(output)
    excel_bytes = output.getvalue()
    output.close()
    return excel_bytes


def make_data_table(df, is_roi_bo=False):
    """Tạo DataTable hiển thị danh sách khách hàng"""
    if df.empty:
        return html.Div("Không có dữ liệu", style={"textAlign": "center", "padding": "20px", "color": "#64748B"})
        
    df_display = df.copy()
    if is_roi_bo:
        columns = [
            {"name": "Cụm", "id": "ten_cum"},
            {"name": "Xã / Bưu cục", "id": "ten_bdx"},
            {"name": "Mã CMS", "id": "cms"},
            {"name": "SL gần nhất", "id": "sl_gan_nhat"},
            {"name": "DT gần nhất", "id": "dt_gan_nhat_disp"},
            {"name": "Kỳ gần nhất có DT", "id": "thang_gan_nhat"}
        ]
        df_display["dt_gan_nhat_disp"] = df_display["dt_gan_nhat"].apply(lambda x: f"{x:,.0f} đ" if pd.notna(x) else "0 đ")
    else:
        columns = [
            {"name": "Cụm", "id": "ten_cum"},
            {"name": "Xã / Bưu cục", "id": "ten_bdx"},
            {"name": "Mã CMS", "id": "cms"},
            {"name": "SL kỳ này", "id": "sl_ky_nay"},
            {"name": "DT kỳ này", "id": "dt_ky_nay_disp"},
            {"name": "SL kỳ trước", "id": "sl_ky_truoc"},
            {"name": "DT kỳ trước", "id": "dt_ky_truoc_disp"},
            {"name": "Chênh lệch DT", "id": "chenh_lech_disp"}
        ]
        df_display["dt_ky_nay_disp"] = df_display["dt_ky_nay"].apply(lambda x: f"{x:,.0f} đ" if pd.notna(x) else "0 đ")
        df_display["dt_ky_truoc_disp"] = df_display["dt_ky_truoc"].apply(lambda x: f"{x:,.0f} đ" if pd.notna(x) else "0 đ")
        df_display["chenh_lech_disp"] = df_display["chenh_lech"].apply(lambda x: f"{x:+,.0f} đ" if pd.notna(x) else "0 đ")

    return dash_table.DataTable(
        columns=columns,
        data=df_display.to_dict('records'),
        sort_action='native',
        filter_action='native',
        page_size=10,
        style_table={'overflowX': 'auto', 'minWidth': '100%'},
        style_header={
            'backgroundColor': '#F8FAFC',
            'fontWeight': 'bold',
            'color': '#1E293B',
            'border': '1px solid #E2E8F0',
            'textAlign': 'left',
            'fontSize': '12px',
            'padding': '10px'
        },
        style_cell={
            'border': '1px solid #E2E8F0',
            'padding': '10px',
            'fontSize': '12px',
            'color': '#334155',
            'fontFamily': 'Inter, sans-serif'
        },
        style_data_conditional=[
            {
                'if': {'row_index': 'odd'},
                'backgroundColor': '#F8FAFC'
            }
        ]
    )


def register_retention_callbacks(app):
    """Đăng ký các callback cho trang Retention v2.0"""

    @app.callback(
        [
            # KPI Cards
            Output("ret-kpi-tang-value", "children"), Output("ret-kpi-tang-subtext", "children"),
            Output("ret-kpi-giam-value", "children"), Output("ret-kpi-giam-subtext", "children"),
            Output("ret-kpi-roibo-value", "children"), Output("ret-kpi-roibo-subtext", "children"),
            Output("ret-kpi-duytri-value", "children"), Output("ret-kpi-duytri-subtext", "children"),
            # Tables
            Output("ret-table-tang-container", "children"),
            Output("ret-table-giam-container", "children"),
            Output("ret-table-roibo-container", "children")
        ],
        [
            Input("btn-apply-filter", "n_clicks"),
            Input("tabs-navigation", "value")
        ],
        [
            State("sidebar-year", "value"),
            State("sidebar-month-select", "value"),
            State("sidebar-week-select", "value"),
            State("sidebar-cycle", "value"),
            State("sidebar-cum", "value")
        ]
    )
    def update_retention_page(n_clicks, tab_val, year, month, week, cycle, cum_val):
        if tab_val is None or tab_val != "tab-retention":
            return [dash.no_update] * 11
        if not year:
            return [dash.no_update] * 11
            
        db_str = str(DB_PATH)
        
        # 1. Gọi hàm analytics lấy biến động dựa trên chu kỳ lọc
        if cycle == 'Tuần':
            if not week:
                return [dash.no_update] * 11
            res = get_weekly_changes(db_str, year, week, cum_val)
        else:
            if not month:
                return [dash.no_update] * 11
            res = get_khhh_changes_v2(db_str, year, month, cum_val)
            
        df_tang = pd.DataFrame(res['tang'])
        df_giam = pd.DataFrame(res['giam'])
        df_roi_bo = pd.DataFrame(res['roi_bo'])
        
        # 2. Tính toán giá trị KPI Cards
        # Nhóm Tăng
        count_tang = len(df_tang)
        sum_tang = df_tang['chenh_lech'].sum() if not df_tang.empty else 0.0
        val_tang = f"{count_tang:,} KH"
        sub_tang = f"+{format_revenue(sum_tang)}"
        
        # Nhóm Giảm
        count_giam = len(df_giam)
        sum_giam = df_giam['chenh_lech'].sum() if not df_giam.empty else 0.0
        val_giam = f"{count_giam:,} KH"
        sub_giam = f"-{format_revenue(abs(sum_giam))}" if sum_giam < 0 else f"{format_revenue(sum_giam)}"
        
        # Nhóm Rời bỏ
        count_roibo = len(df_roi_bo)
        sum_roibo = df_roi_bo['dt_gan_nhat'].sum() if not df_roi_bo.empty else 0.0
        val_roibo = f"{count_roibo:,} KH"
        sub_roibo = f"-{format_revenue(sum_roibo)}"
        
        # Nhóm Duy trì
        val_duytri = f"{res['duy_tri_count']:,} KH"
        sub_duytri = "Doanh thu không đổi và > 0"
        
        # 3. Tạo DataTable cho 3 bảng chi tiết
        table_tang = make_data_table(df_tang, is_roi_bo=False)
        table_giam = make_data_table(df_giam, is_roi_bo=False)
        table_roibo = make_data_table(df_roi_bo, is_roi_bo=True)
        
        return [
            val_tang, sub_tang,
            val_giam, sub_giam,
            val_roibo, sub_roibo,
            val_duytri, sub_duytri,
            table_tang, table_giam, table_roibo
        ]

    # Callback xuất Excel nhóm TĂNG
    @app.callback(
        Output("ret-download-tang", "data"),
        Input("ret-btn-export-tang", "n_clicks"),
        [
            State("tabs-navigation", "value"),
            State("sidebar-year", "value"),
            State("sidebar-month-select", "value"),
            State("sidebar-week-select", "value"),
            State("sidebar-cycle", "value"),
            State("sidebar-cum", "value")
        ],
        prevent_initial_call=True
    )
    def export_tang_excel(n_clicks, tab_val, year, month, week, cycle, cum_val):
        if not n_clicks or tab_val != "tab-retention" or not year:
            return dash.no_update
            
        db_str = str(DB_PATH)
        if cycle == 'Tuần':
            if not week: return dash.no_update
            res = get_weekly_changes(db_str, year, week, cum_val)
            period_str = f"Tuan {week}_{year}"
        else:
            if not month: return dash.no_update
            res = get_khhh_changes_v2(db_str, year, month, cum_val)
            period_str = f"Thang {month:02d}_{year}"
            
        df = pd.DataFrame(res['tang'])
        if df.empty:
            return dash.no_update
            
        excel_bytes = generate_retention_list_excel(df, "KH Hien Huu Tang Doanh Thu", False, period_str)
        return dcc.send_bytes(excel_bytes, filename=f"KH_Tang_{period_str}.xlsx")

    # Callback xuất Excel nhóm GIẢM
    @app.callback(
        Output("ret-download-giam", "data"),
        Input("ret-btn-export-giam", "n_clicks"),
        [
            State("tabs-navigation", "value"),
            State("sidebar-year", "value"),
            State("sidebar-month-select", "value"),
            State("sidebar-week-select", "value"),
            State("sidebar-cycle", "value"),
            State("sidebar-cum", "value")
        ],
        prevent_initial_call=True
    )
    def export_giam_excel(n_clicks, tab_val, year, month, week, cycle, cum_val):
        if not n_clicks or tab_val != "tab-retention" or not year:
            return dash.no_update
            
        db_str = str(DB_PATH)
        if cycle == 'Tuần':
            if not week: return dash.no_update
            res = get_weekly_changes(db_str, year, week, cum_val)
            period_str = f"Tuan {week}_{year}"
        else:
            if not month: return dash.no_update
            res = get_khhh_changes_v2(db_str, year, month, cum_val)
            period_str = f"Thang {month:02d}_{year}"
            
        df = pd.DataFrame(res['giam'])
        if df.empty:
            return dash.no_update
            
        excel_bytes = generate_retention_list_excel(df, "KH Hien Huu Giam Doanh Thu", False, period_str)
        return dcc.send_bytes(excel_bytes, filename=f"KH_Giam_{period_str}.xlsx")

    # Callback xuất Excel nhóm RỜI BỎ
    @app.callback(
        Output("ret-download-roibo", "data"),
        Input("ret-btn-export-roibo", "n_clicks"),
        [
            State("tabs-navigation", "value"),
            State("sidebar-year", "value"),
            State("sidebar-month-select", "value"),
            State("sidebar-week-select", "value"),
            State("sidebar-cycle", "value"),
            State("sidebar-cum", "value")
        ],
        prevent_initial_call=True
    )
    def export_roibo_excel(n_clicks, tab_val, year, month, week, cycle, cum_val):
        if not n_clicks or tab_val != "tab-retention" or not year:
            return dash.no_update
            
        db_str = str(DB_PATH)
        if cycle == 'Tuần':
            if not week: return dash.no_update
            res = get_weekly_changes(db_str, year, week, cum_val)
            period_str = f"Tuan {week}_{year}"
        else:
            if not month: return dash.no_update
            res = get_khhh_changes_v2(db_str, year, month, cum_val)
            period_str = f"Thang {month:02d}_{year}"
            
        df = pd.DataFrame(res['roi_bo'])
        if df.empty:
            return dash.no_update
            
        excel_bytes = generate_retention_list_excel(df, "KH Hien Huu Roi Bo Churn", True, period_str)
        return dcc.send_bytes(excel_bytes, filename=f"KH_Roibo_{period_str}.xlsx")
