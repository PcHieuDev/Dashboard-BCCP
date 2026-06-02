# -*- coding: utf-8 -*-
"""
Các hàm helper hỗ trợ xuất báo cáo doanh thu ra định dạng Excel (openpyxl) và PDF (reportlab).
Hỗ trợ Unicode tiếng Việt đầy đủ và format số liệu đẹp mắt.
"""

import io
import os
import pandas as pd
from datetime import datetime

# Import thư viện Excel
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Import thư viện PDF
# Bỏ reportlab theo yêu cầu dọn dẹp PDF

from components.data_table import COL_LABEL_MAP

# --------------------------------------------------------------------------
# ĐĂNG KÝ FONT TIẾNG VIỆT CHO PDF
# --------------------------------------------------------------------------
# Đăng ký font tiếng Việt cho PDF - Bỏ theo yêu cầu dọn dẹp PDF
FONT_NAME = 'Helvetica'


# --------------------------------------------------------------------------
# HÀM TẠO EXCEL
# --------------------------------------------------------------------------
def generate_excel_report(df: pd.DataFrame, groupby_cols: list, compare_opt: str, filter_info: dict) -> bytes:
    """
    Tạo báo cáo Excel chuyên nghiệp từ DataFrame và thông tin bộ lọc.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Báo cáo Doanh thu"
    
    font_name = "Arial"
    
    # 1. Tiêu đề báo cáo
    ws.merge_cells("A1:H1")
    ws["A1"] = "BÁO CÁO DOANH THU ĐIỀU HÀNH BCCP"
    ws["A1"].font = Font(name=font_name, size=16, bold=True, color="003366")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40
    
    # 2. Thông tin bộ lọc
    ws["A3"] = "Thông tin bộ lọc:"
    ws["A3"].font = Font(name=font_name, size=11, bold=True, italic=True, color="555555")
    
    row_idx = 4
    for k, v in filter_info.items():
        ws.cell(row=row_idx, column=1, value=f"• {k}: {v}").font = Font(name=font_name, size=10, italic=True)
        row_idx += 1
        
    ws.cell(row=row_idx, column=1, value=f"• Ngày xuất báo cáo: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}").font = Font(name=font_name, size=10, italic=True)
    row_idx += 2  # Để trống 1 dòng
    
    # 3. Định nghĩa các cột cần xuất
    columns = []
    for col in groupby_cols:
        columns.append(col)
    
    metrics = ['san_luong', 'khoi_luong_thuc', 'cuoc_cb_tong', 'cuoc_tt_tong', 'cuoc_tt_gom_vat', 'so_kh']
    for col in metrics:
        columns.append(col)
        
    if compare_opt in ('prev_period', 'both'):
        for m in metrics:
            columns.append(f"{m}_prev")
            columns.append(f"{m}_pct_change")
            
    if compare_opt in ('yoy', 'both'):
        for m in metrics:
            columns.append(f"{m}_yoy")
            columns.append(f"{m}_yoy_pct_change")
            
    # 4. Vẽ header bảng dữ liệu
    header_row = row_idx
    ws.row_dimensions[header_row].height = 28
    
    header_fill = PatternFill(start_color="0056B3", end_color="0056B3", fill_type="solid") # Màu xanh dương bưu điện
    header_font = Font(name=font_name, size=10, bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border_thin = Side(border_style="thin", color="CBD5E1")
    cell_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=COL_LABEL_MAP.get(col_name, col_name))
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = cell_border
        
    # 5. Điền dữ liệu
    data_start_row = header_row + 1
    for r_idx, row_data in enumerate(df.to_dict('records'), start=data_start_row):
        ws.row_dimensions[r_idx].height = 20
        # Xen kẽ màu dòng
        is_odd = (r_idx % 2 == 1)
        row_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid") if is_odd else PatternFill(fill_type=None)
        
        for c_idx, col_name in enumerate(columns, start=1):
            val = row_data.get(col_name, None)
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.border = cell_border
            if row_fill.fill_type:
                cell.fill = row_fill
                
            # Phân loại format
            if pd.isna(val) or val is None:
                cell.value = "-"
                cell.alignment = Alignment(horizontal="center")
            elif col_name in groupby_cols:
                cell.value = str(val)
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.font = Font(name=font_name, size=9, bold=True)
            elif 'pct_change' in col_name:
                # Đổi thành tỉ lệ thực tế để Excel tự nhân 100 khi định dạng %
                cell.value = float(val) / 100.0
                cell.number_format = '+0.0%;-0.0%;"0.0%"'
                cell.alignment = Alignment(horizontal="right", vertical="center")
                # Màu đỏ nếu âm, xanh nếu dương
                if val < 0:
                    cell.font = Font(name=font_name, size=9, color="DC2626")
                elif val > 0:
                    cell.font = Font(name=font_name, size=9, color="16A34A")
                else:
                    cell.font = Font(name=font_name, size=9)
            elif col_name in ('san_luong', 'san_luong_prev', 'san_luong_yoy', 'so_kh', 'so_kh_prev', 'so_kh_yoy'):
                cell.value = int(val)
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.font = Font(name=font_name, size=9)
            elif 'khoi_luong_thuc' in col_name:
                # Chuyển đổi gram -> kg
                cell.value = float(val) / 1000.0
                cell.number_format = '#,##0.0" kg"'
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.font = Font(name=font_name, size=9)
            elif 'cuoc_' in col_name:
                cell.value = float(val)
                cell.number_format = '#,##0" đ"'
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.font = Font(name=font_name, size=9)
            else:
                cell.value = val
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
    # 6. Thêm dòng tổng cộng ở cuối
    total_row = data_start_row + len(df)
    ws.row_dimensions[total_row].height = 22
    total_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    
    # Ghi nhãn "Tổng cộng" ở cột 1
    cell_tot_lbl = ws.cell(row=total_row, column=1, value="Tổng cộng")
    cell_tot_lbl.font = Font(name=font_name, size=10, bold=True, color="000000")
    cell_tot_lbl.alignment = Alignment(horizontal="left", vertical="center")
    cell_tot_lbl.fill = total_fill
    cell_tot_lbl.border = cell_border
    
    # Tô nền các cột groupby còn lại (nếu có)
    for c_idx in range(2, len(groupby_cols) + 1):
        cell_empty = ws.cell(row=total_row, column=c_idx, value="")
        cell_empty.fill = total_fill
        cell_empty.border = cell_border
        
    # Tính tổng cho từng cột metrics
    for c_idx, col_name in enumerate(columns[len(groupby_cols):], start=len(groupby_cols) + 1):
        cell = ws.cell(row=total_row, column=c_idx)
        cell.fill = total_fill
        cell.border = cell_border
        cell.font = Font(name=font_name, size=9, bold=True)
        
        # Chỉ tính SUM cho các cột số lượng, cước, khối lượng.
        # Với cột phần trăm biến động, ta sẽ tính lại dựa trên tổng.
        if 'pct_change' in col_name:
            # Rút gọn: tính trung bình hoặc để trống
            cell.value = "-"
            cell.alignment = Alignment(horizontal="center")
        elif col_name in ('so_kh', 'so_kh_prev', 'so_kh_yoy'):
            # Số lượng khách hàng tổng: lấy sum (gần đúng)
            col_letter = get_column_letter(c_idx)
            cell.value = f"=SUM({col_letter}{data_start_row}:{col_letter}{total_row-1})"
            cell.number_format = '#,##0'
            cell.alignment = Alignment(horizontal="right")
        elif 'khoi_luong_thuc' in col_name:
            col_letter = get_column_letter(c_idx)
            cell.value = f"=SUM({col_letter}{data_start_row}:{col_letter}{total_row-1})"
            cell.number_format = '#,##0.0" kg"'
            cell.alignment = Alignment(horizontal="right")
        elif col_name in ('san_luong', 'san_luong_prev', 'san_luong_yoy') or 'cuoc_' in col_name:
            col_letter = get_column_letter(c_idx)
            cell.value = f"=SUM({col_letter}{data_start_row}:{col_letter}{total_row-1})"
            if 'cuoc_' in col_name:
                cell.number_format = '#,##0" đ"'
            else:
                cell.number_format = '#,##0'
            cell.alignment = Alignment(horizontal="right")
            
    # 7. Tự động giãn cột phù hợp dữ liệu
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        
        # Kiểm tra độ dài tiêu đề hiển thị
        for cell in col:
            # Bỏ qua hàng tiêu đề chính dòng 1 đã merge
            if cell.row == 1:
                continue
            val_str = str(cell.value or '')
            # Tối ưu đo độ dài
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    # Xuất dữ liệu nhị phân
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


# --------------------------------------------------------------------------
# HÀM TẠO PDF - ĐÃ XÓA
# --------------------------------------------------------------------------


def generate_customer_excel(df: pd.DataFrame, filter_info_str: str) -> bytes:
    """
    Tạo báo cáo Excel chuyên nghiệp với header 2 tầng và dòng tổng cộng cho dữ liệu Chi tiết Khách hàng (CMS).
    """
    import re
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Chi tiết Khách hàng"
    
    font_name = "Arial"
    
    # 1. Tiêu đề báo cáo
    ws.merge_cells("A1:H1")
    ws["A1"] = "BÁO CÁO CHI TIẾT DOANH THU THEO KHÁCH HÀNG (CMS)"
    ws["A1"].font = Font(name=font_name, size=14, bold=True, color="003366")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40
    
    # 2. Thông tin bộ lọc
    ws["A3"] = "Thông tin bộ lọc:"
    ws["A3"].font = Font(name=font_name, size=11, bold=True, italic=True, color="555555")
    
    ws["A4"] = f"• Bộ lọc áp dụng: {filter_info_str}"
    ws["A4"].font = Font(name=font_name, size=10, italic=True)
    
    ws["A5"] = f"• Ngày xuất báo cáo: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    ws["A5"].font = Font(name=font_name, size=10, italic=True)
    
    # 3. Phân tích các cột
    base_cols = ['cms', 'loai_kh', 'hop_dong', 'buu_cuc_list']
    pivot_cols = [c for c in df.columns if c not in base_cols and not c.startswith('Tổng')]
    total_cols = [c for c in ['Tổng SL', 'Tổng KL', 'Tổng Cước TT', 'Tổng Cước VAT'] if c in df.columns]
    
    columns = base_cols + pivot_cols + total_cols
    num_cols = len(columns)
    
    # 4. Thiết lập Header 2 tầng ở hàng 7 và 8
    row_1 = 7
    row_2 = 8
    
    ws.row_dimensions[row_1].height = 25
    ws.row_dimensions[row_2].height = 25
    
    col_labels = {
        'cms': 'Mã khách hàng',
        'loai_kh': 'Loại khách hàng',
        'hop_dong': 'Trạng thái HĐ',
        'buu_cuc_list': 'Danh sách Bưu cục',
    }
    
    # Cột cơ bản: Gộp đứng
    for c_idx in range(1, 5):
        col_letter = get_column_letter(c_idx)
        ws.merge_cells(f"{col_letter}{row_1}:{col_letter}{row_2}")
        col_name = columns[c_idx - 1]
        ws.cell(row=row_1, column=c_idx, value=col_labels.get(col_name, col_name))
        
    # Cột dịch vụ: Gộp ngang theo nhóm dịch vụ
    def get_group_name(col):
        match = re.match(r'^\[(.*?)\]', col)
        return match.group(1) if match else "Khác"
        
    current_g_name = None
    g_start_idx = None
    
    for idx, col in enumerate(pivot_cols, start=5):
        g_name = get_group_name(col)
        if g_name != current_g_name:
            if current_g_name is not None:
                start_letter = get_column_letter(g_start_idx)
                end_letter = get_column_letter(idx - 1)
                ws.merge_cells(f"{start_letter}{row_1}:{end_letter}{row_1}")
                ws.cell(row=row_1, column=g_start_idx, value=current_g_name)
            current_g_name = g_name
            g_start_idx = idx
            
        metric_short = col.split(']')[-1].strip()
        metric_label_map = {
            'SL': 'Sản lượng (SL)',
            'KL': 'Khối lượng (KL-kg)',
            'Cước TT': 'Cước TT',
            'Cước VAT': 'Cước gồm VAT'
        }
        ws.cell(row=row_2, column=idx, value=metric_label_map.get(metric_short, metric_short))
        
    # Merge nhóm dịch vụ cuối cùng
    if current_g_name is not None and g_start_idx is not None:
        start_letter = get_column_letter(g_start_idx)
        end_letter = get_column_letter(4 + len(pivot_cols))
        ws.merge_cells(f"{start_letter}{row_1}:{end_letter}{row_1}")
        ws.cell(row=row_1, column=g_start_idx, value=current_g_name)
        
    # Cột Tổng cộng: Gộp ngang
    if total_cols:
        start_tot_idx = 5 + len(pivot_cols)
        end_tot_idx = start_tot_idx + len(total_cols) - 1
        start_letter = get_column_letter(start_tot_idx)
        end_letter = get_column_letter(end_tot_idx)
        ws.merge_cells(f"{start_letter}{row_1}:{end_letter}{row_1}")
        ws.cell(row=row_1, column=start_tot_idx, value="Tổng cộng")
        
        total_label_map = {
            'Tổng SL': 'Tổng Sản lượng',
            'Tổng KL': 'Tổng Khối lượng (kg)',
            'Tổng Cước TT': 'Tổng Cước TT',
            'Tổng Cước VAT': 'Tổng Cước gồm VAT'
        }
        for idx, col in enumerate(total_cols, start=start_tot_idx):
            ws.cell(row=row_2, column=idx, value=total_label_map.get(col, col))
            
    # Áp dụng Style cho Header
    header_fill = PatternFill(start_color="0056B3", end_color="0056B3", fill_type="solid")
    header_font = Font(name=font_name, size=10, bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border_thin = Side(border_style="thin", color="CBD5E1")
    cell_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    
    for col_idx in range(1, num_cols + 1):
        cell1 = ws.cell(row=row_1, column=col_idx)
        cell2 = ws.cell(row=row_2, column=col_idx)
        for cell in (cell1, cell2):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = cell_border
            
    # 5. Điền dữ liệu
    data_start_row = row_2 + 1
    for r_idx, row_data in enumerate(df.to_dict('records'), start=data_start_row):
        ws.row_dimensions[r_idx].height = 20
        is_odd = (r_idx % 2 == 1)
        row_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid") if is_odd else PatternFill(fill_type=None)
        
        for c_idx, col_name in enumerate(columns, start=1):
            val = row_data.get(col_name, None)
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.border = cell_border
            if row_fill.fill_type:
                cell.fill = row_fill
                
            if pd.isna(val) or val is None:
                cell.value = "-"
                cell.alignment = Alignment(horizontal="center")
            elif col_name == 'cms':
                cell.value = str(val)
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.font = Font(name=font_name, size=9, bold=True)
            elif col_name in ['loai_kh', 'hop_dong']:
                cell.value = str(val)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(name=font_name, size=9)
            elif col_name == 'buu_cuc_list':
                cell.value = str(val)
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.font = Font(name=font_name, size=9)
            elif 'SL' in col_name or 'Tổng SL' == col_name:
                cell.value = int(val)
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.font = Font(name=font_name, size=9)
            elif 'KL' in col_name or 'Tổng KL' == col_name:
                cell.value = float(val)
                cell.number_format = '#,##0.0'
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.font = Font(name=font_name, size=9)
            elif 'Cước' in col_name or 'Tổng Cước' in col_name:
                cell.value = float(val)
                cell.number_format = '#,##0" đ"'
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.font = Font(name=font_name, size=9)
            else:
                cell.value = val
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
    # 6. Dòng tổng cộng ở cuối
    total_row = data_start_row + len(df)
    ws.row_dimensions[total_row].height = 22
    total_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    
    ws.merge_cells(f"A{total_row}:D{total_row}")
    cell_tot_lbl = ws.cell(row=total_row, column=1, value="Tổng cộng")
    cell_tot_lbl.font = Font(name=font_name, size=10, bold=True, color="000000")
    cell_tot_lbl.alignment = Alignment(horizontal="left", vertical="center")
    
    for c_idx in range(1, num_cols + 1):
        cell = ws.cell(row=total_row, column=c_idx)
        cell.fill = total_fill
        cell.border = cell_border
        if c_idx > 4:
            cell.font = Font(name=font_name, size=9, bold=True)
            col_letter = get_column_letter(c_idx)
            cell.value = f"=SUM({col_letter}{data_start_row}:{col_letter}{total_row - 1})"
            cell.alignment = Alignment(horizontal="right", vertical="center")
            
            col_name = columns[c_idx - 1]
            if 'SL' in col_name or 'Tổng SL' == col_name:
                cell.number_format = '#,##0'
            elif 'KL' in col_name or 'Tổng KL' == col_name:
                cell.number_format = '#,##0.0'
            elif 'Cước' in col_name or 'Tổng Cước' in col_name:
                cell.number_format = '#,##0" đ"'
                
    # 7. Tự động điều chỉnh độ rộng cột
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row in [1, 2, 3, 4, 5, 6, 7, 8]:
                continue
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
