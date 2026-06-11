# -*- coding: utf-8 -*-
"""
Callbacks xử lý dữ liệu trang Khách hàng mới/tái bán v2.0 (/bccp/new-customer).
Tải dữ liệu từ new_customers và transactions để tính toán 3 chỉ số KPI và bảng chi tiết xã.
"""

import sqlite3
import pandas as pd
import dash
from dash import Output, Input, State, html, dcc, dash_table
import dash_bootstrap_components as dbc
from datetime import datetime
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

import sys
from pathlib import Path

# Thêm thư mục gốc vào sys.path để import
project_root = Path(__file__).resolve().parent.parent.parent
if str(project_root) not in sys.path:
    sys.path.insert(0, str(project_root))

from config.settings import DB_PATH
from callbacks.utils import format_revenue

import logging
try:
    from config.logger import get_logger
    logger = get_logger(__name__)
except ImportError:
    import sys
    from pathlib import Path
    sys.path.append(str(Path(__file__).resolve().parent.parent.parent))
    try:
        from config.logger import get_logger
        logger = get_logger(__name__)
    except ImportError:
        logger = logging.getLogger(__name__)


def get_4_months_range(year, month):
    """Lấy danh sách 4 cặp (thang, nam) gần nhất bao gồm cả tháng hiện tại"""
    res = []
    curr_m, curr_y = month, year
    for _ in range(4):
        res.append((curr_m, curr_y))
        curr_m -= 1
        if curr_m == 0:
            curr_m = 12
            curr_y -= 1
    return res

def query_and_process_new_customers(year, month, week, cycle, cum_val, bdx_val, bc_val):
    """
    Xử lý logic tính KPI và bảng cho KH mới/tái bán.
    - Lọc Tháng:
      1. KPI 1: Số KH mới phát sinh trong tháng = count(new_customers) của tháng target
      2. KPI 2: Tổng KH mới 4 tháng gần nhất = count(new_customers) của 4 tháng
      3. KPI 3: Tổng doanh thu tháng hiện tại của KH 4 tháng gần nhất
    - Lọc Tuần:
      1. KPI 1: Số KH mới 4 tháng có phát sinh doanh thu trong tuần target
      2. KPI 2: Giữ nguyên (logic tháng)
      3. KPI 3: Tổng doanh thu trong tuần của KH 4 tháng gần nhất
    """
    if not DB_PATH.exists():
        return pd.DataFrame(), 0, 0, 0.0
        
    conn = sqlite3.connect(str(DB_PATH))
    cursor = conn.cursor()
    
    import config.week_calendar as calendar_helper
    
    try:
        # A. Xác định 4 tháng gần nhất
        months_4 = get_4_months_range(year, month if cycle == 'Tháng' else datetime.now().month)
        # Tạo chuỗi query cho 4 tháng
        placeholders_4m = " OR ".join(["(thang = ? AND nam = ?)" for _ in months_4])
        params_4m = []
        for m, y in months_4:
            params_4m.extend([m, y])
            
        # B. Lọc địa lý
        geo_where = []
        geo_params = []
        if cum_val and cum_val != "Tất cả":
            geo_where.append("ten_cum = ?")
            geo_params.append(cum_val)
        if bdx_val and bdx_val != "Tất cả":
            # Xã/Phường
            geo_where.append("ma_bdx = (SELECT DISTINCT ma_bdx FROM dim_buucuc WHERE ten_bdx = ? LIMIT 1)")
            geo_params.append(bdx_val)
        if bc_val and bc_val != "Tất cả":
            geo_where.append("buu_cuc = ?")
            geo_params.append(bc_val)
            
        geo_where_str = " AND " + " AND ".join(geo_where) if geo_where else ""
        
        # 1. Tính toán KH 4 tháng gần nhất
        q_4m = f"""
            SELECT DISTINCT cms, thang, nam, buu_cuc, ma_bdx, ten_cum, nhom_dv, tong_doanh_thu, ngay_phat_sinh 
            FROM new_customers 
            WHERE ({placeholders_4m}) {geo_where_str}
        """
        df_4m = pd.read_sql_query(q_4m, conn, params=params_4m + geo_params)
        kpi_4m_count = df_4m['cms'].nunique()
        
        # 2. Xử lý theo Lọc Tháng / Lọc Tuần
        if cycle == 'Tháng':
            # KPI 1: Số KH mới tháng hiện tại
            df_curr = df_4m[(df_4m['thang'] == month) & (df_4m['nam'] == year)]
            kpi_curr_count = df_curr['cms'].nunique()
            
            # KPI 3: Tổng doanh thu tháng hiện tại của KH 4 tháng gần nhất
            # Lấy doanh thu từ bảng transactions của tháng hiện tại cho các CMS trong df_4m
            kpi_revenue = 0.0
            if not df_4m.empty:
                cms_list = df_4m['cms'].unique().tolist()
                placeholders_cms = ",".join(["?"] * len(cms_list))
                q_rev = f"""
                    SELECT SUM(cuoc_tt_tong) 
                    FROM transactions 
                    WHERE nam_du_lieu = ? AND thang_du_lieu = ? AND cms IN ({placeholders_cms})
                """
                cursor.execute(q_rev, [year, f"T{month:02d}"] + cms_list)
                kpi_revenue = cursor.fetchone()[0] or 0.0
                
            # Bảng chi tiết xã (Khi lọc tháng: hiện KH mới của tháng đó)
            df_table_data = df_curr.copy()
            
        else: # Tuần
            # Xác định ngày bắt đầu và kết thúc của tuần target
            weeks_list = calendar_helper.get_week_list(year)
            w_start, w_end = None, None
            for w_num, start_d, end_d in weeks_list:
                if w_num == week:
                    w_start, w_end = start_d.isoformat(), end_d.isoformat()
                    break
                    
            if not w_start:
                w_start, w_end = f"{year}-01-01", f"{year}-12-31"
                
            # Tìm danh sách CMS 4 tháng có phát sinh doanh thu trong tuần này
            kpi_curr_count = 0
            kpi_revenue = 0.0
            df_table_data = pd.DataFrame()
            
            if not df_4m.empty:
                cms_list = df_4m['cms'].unique().tolist()
                placeholders_cms = ",".join(["?"] * len(cms_list))
                
                # Query doanh thu và số lượng KH tuần
                q_week_tx = f"""
                    SELECT t.cms, t.buu_cuc, SUM(t.cuoc_tt_tong) as dt, MIN(t.ngay_chap_nhan) as ngay_phat_sinh
                    FROM transactions t
                    WHERE t.ngay_chap_nhan BETWEEN ? AND ? 
                      AND t.cms IN ({placeholders_cms})
                    GROUP BY t.cms, t.buu_cuc
                """
                df_week_tx = pd.read_sql_query(q_week_tx, conn, params=[w_start, w_end] + cms_list)
                
                if not df_week_tx.empty:
                    # Lọc doanh thu > 0
                    df_week_tx = df_week_tx[df_week_tx['dt'] > 0]
                    kpi_curr_count = df_week_tx['cms'].nunique()
                    kpi_revenue = df_week_tx['dt'].sum()
                    
                    # Merge thông tin địa lý và nhóm dịch vụ con từ df_4m
                    df_geo_map = df_4m[['cms', 'ten_cum', 'ma_bdx', 'nhom_dv']].drop_duplicates(subset=['cms'])
                    df_table_data = pd.merge(df_week_tx, df_geo_map, on='cms', how='inner')
                    df_table_data = df_table_data.rename(columns={'dt': 'tong_doanh_thu'})
                    
        # 3. Hoàn thiện bảng dữ liệu chi tiết xã
        # Bổ sung Tên Xã từ dim_buucuc
        if not df_table_data.empty:
            df_buucuc = pd.read_sql_query("SELECT ma_bc as buu_cuc, ten_bdx FROM dim_buucuc", conn)
            df_table_data = pd.merge(df_table_data, df_buucuc, on='buu_cuc', how='left')
            # Cột mong đợi: ten_cum | ten_bdx | buu_cuc | cms | ngay_phat_sinh | nhom_dv | tong_doanh_thu
            # Điền fallback nếu thiếu nhom_dv
            if 'nhom_dv' not in df_table_data.columns:
                df_table_data['nhom_dv'] = 'BCCP'
            df_table_data = df_table_data[['ten_cum', 'ten_bdx', 'buu_cuc', 'cms', 'ngay_phat_sinh', 'nhom_dv', 'tong_doanh_thu']]
            # Sắp xếp giảm dần theo doanh thu
            df_table_data = df_table_data.sort_values(by='tong_doanh_thu', ascending=False)
        else:
            df_table_data = pd.DataFrame(columns=['ten_cum', 'ten_bdx', 'buu_cuc', 'cms', 'ngay_phat_sinh', 'nhom_dv', 'tong_doanh_thu'])
            
        return df_table_data, kpi_curr_count, kpi_4m_count, kpi_revenue
        
    except Exception as e:
        logger.error(f"Lỗi query KH mới: {e}")
        import traceback
        traceback.print_exc()
        return pd.DataFrame(), 0, 0, 0.0
    finally:
        conn.close()

def register_new_customer_callbacks(app):
    """Đăng ký callbacks cho trang KH mới v2.0"""
    
    @app.callback(
        [
            Output("new-cust-kpi-count-value", "children"),
            Output("new-cust-kpi-count-subtext", "children"),
            Output("new-cust-kpi-4m-count-value", "children"),
            Output("new-cust-kpi-4m-count-subtext", "children"),
            Output("new-cust-kpi-revenue-value", "children"),
            Output("new-cust-kpi-revenue-subtext", "children"),
            Output("new-cust-table-container", "children")
        ],
        [Input("btn-apply-filter", "n_clicks")],
        [
            State("sidebar-year", "value"),
            State("sidebar-month-select", "value"),
            State("sidebar-week-select", "value"),
            State("sidebar-period", "value"),
            State("sidebar-cum", "value"),
            State("sidebar-bdx", "value"),
            State("sidebar-buu-cuc", "value")
        ]
    )
    def update_new_cust_dashboard(n_clicks, year, month, week, cycle, cum_val, bdx_val, bc_val):
        if not year or (cycle == 'Tháng' and not month) or (cycle == 'Tuần' and not week):
            return "—", "Số KH mới phát sinh trong kỳ", "—", "Tổng KH mới 4 tháng gần nhất", "—", "Doanh thu bán mới tích lũy", html.Div("Vui lòng chọn bộ lọc.")
            
        df, kpi_curr, kpi_4m, kpi_rev = query_and_process_new_customers(year, month, week, cycle, cum_val, bdx_val, bc_val)
        
        # Format strings
        curr_label = f"KH phát sinh trong Tháng {month:02d}/{year}" if cycle == 'Tháng' else f"KH có doanh thu Tuần {week}/{year}"
        curr_sub = f"Theo định nghĩa doanh thu > 0 ({curr_label})"
        
        kpi_4m_sub = f"Thời kỳ 4 tháng gần nhất tính đến nay"
        kpi_rev_sub = f"Phát sinh trong kỳ (Tháng {month:02d} hoặc Tuần {week}) của KH 4 tháng"
        
        # Render DataTable
        if df.empty:
            table_layout = html.Div("Không tìm thấy khách hàng mới phát sinh doanh thu trong kỳ này.", style={"color": "#64748B", "padding": "20px", "textAlign": "center"})
        else:
            df_display = df.copy()
            df_display['tong_doanh_thu'] = df_display['tong_doanh_thu'].map(lambda x: f"{x:,.0f} đ" if x > 0 else "0 đ")
            
            columns = [
                {"name": "Cụm địa lý", "id": "ten_cum"},
                {"name": "Phường / Xã", "id": "ten_bdx"},
                {"name": "Mã Bưu Cục", "id": "buu_cuc"},
                {"name": "Mã CMS", "id": "cms"},
                {"name": "Ngày phát sinh", "id": "ngay_phat_sinh"},
                {"name": "Nhóm DV chính", "id": "nhom_dv"},
                {"name": "Doanh Thu", "id": "tong_doanh_thu"}
            ]
            
            table_layout = dash_table.DataTable(
                data=df_display.to_dict('records'),
                columns=columns,
                page_size=15,
                sort_action="native",
                filter_action="native",
                style_table={'overflowX': 'auto', 'borderRadius': '8px', 'border': '1px solid #E2E8F0'},
                style_header={
                    'backgroundColor': '#F8FAFC',
                    'fontWeight': 'bold',
                    'color': '#1E293B',
                    'border': '1px solid #CBD5E1'
                },
                style_cell={
                    'padding': '8px 10px',
                    'textAlign': 'left',
                    'fontSize': '13px',
                    'fontFamily': 'Inter, sans-serif'
                }
            )
            
        return (
            f"{kpi_curr:,} KH", curr_sub,
            f"{kpi_4m:,} KH", kpi_4m_sub,
            format_revenue(kpi_rev), kpi_rev_sub,
            table_layout
        )

    # 4. Callback xuất file Excel
    @app.callback(
        Output("new-cust-download", "data"),
        [Input("new-cust-btn-export-excel", "n_clicks")],
        [
            State("sidebar-year", "value"),
            State("sidebar-month-select", "value"),
            State("sidebar-week-select", "value"),
            State("sidebar-period", "value"),
            State("sidebar-cum", "value"),
            State("sidebar-bdx", "value"),
            State("sidebar-buu-cuc", "value")
        ],
        prevent_initial_call=True
    )
    def export_new_cust_excel(n_clicks, year, month, week, cycle, cum_val, bdx_val, bc_val):
        if not n_clicks:
            return dash.no_update
            
        df, _, _, _ = query_and_process_new_customers(year, month, week, cycle, cum_val, bdx_val, bc_val)
        if df.empty:
            return dash.no_update
            
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "DS_KH_Moi"
        ws.views.sheetView[0].showGridLines = True
        
        # Setup styles
        title_font = Font(name="Arial", size=14, bold=True, color="1E3A8A")
        sub_font = Font(name="Arial", size=10, italic=True)
        header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        data_font = Font(name="Arial", size=10)
        header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )
        
        ws['A1'] = f"DANH SÁCH KHÁCH HÀNG MỚI PHÁT SINH"
        ws['A1'].font = title_font
        
        filter_label = f"Năm {year} | Chu kỳ: {cycle} " + (f"(Tháng {month:02d})" if cycle == 'Tháng' else f"(Tuần {week})")
        ws['A2'] = f"Bộ lọc: {filter_label} | Lọc địa lý: Cụm {cum_val} - BC {bc_val}"
        ws['A2'].font = sub_font
        
        headers = ["Cụm địa lý", "Phường / Xã", "Mã Bưu cục", "Mã CMS", "Ngày phát sinh", "Nhóm DV chính", "Doanh thu"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            
        ws.row_dimensions[4].height = 24
        
        row_num = 4
        for _, r in df.iterrows():
            row_num += 1
            ws.cell(row=row_num, column=1, value=r['ten_cum'])
            ws.cell(row=row_num, column=2, value=r['ten_bdx'])
            ws.cell(row=row_num, column=3, value=r['buu_cuc'])
            ws.cell(row=row_num, column=4, value=r['cms'])
            ws.cell(row=row_num, column=5, value=r['ngay_phat_sinh'])
            ws.cell(row=row_num, column=6, value=r['nhom_dv'])
            
            cell_rev = ws.cell(row=row_num, column=7, value=r['tong_doanh_thu'])
            cell_rev.number_format = '#,##0" đ"'
            cell_rev.alignment = Alignment(horizontal="right")
            
            for col_idx in range(1, 8):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.font = data_font
                cell.border = thin_border
                if col_idx != 7:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 30)
            
        out = io.BytesIO()
        wb.save(out)
        excel_bytes = out.getvalue()
        out.close()
        
        filename = f"Danh_sach_KHM_{cycle}_{month if cycle == 'Tháng' else week}_{year}.xlsx"
        return dcc.send_bytes(excel_bytes, filename=filename)
