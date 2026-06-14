# -*- coding: utf-8 -*-
"""
Script đồng bộ dữ liệu tham chiếu (bưu cục, sản phẩm/dịch vụ) từ database SQLite
vào các sheet Ref của 3 file Excel mẫu nhập liệu, đồng thời nâng cấp toàn bộ công thức
đối chiếu thông minh (tự động xử lý cả định dạng Chuỗi và định dạng Số cho mã bưu cục)
cho tất cả các dòng dữ liệu.
"""

import os
import sys
import sqlite3
import pandas as pd
import openpyxl
from pathlib import Path

# Đảm bảo in tiếng Việt chính xác trên Windows Console
try:
    sys.stdout.reconfigure(encoding='utf-8')
except Exception:
    pass

# Cấu hình đường dẫn
PROJECT_ROOT = Path(__file__).resolve().parent.parent
DB_PATH = Path(r"E:\OneDrive\z.Database-TTKD-Data\dashboard.db")
TEMPLATES_DIR = PROJECT_ROOT / "data" / "mau-file-import"

TEMPLATE_FILES = {
    "dich_vu_khac": TEMPLATES_DIR / "mau_import_dich_vu_khac.xlsx",
    "doanh_thu_bccp": TEMPLATES_DIR / "mau_import_doanh_thu_BCCP.xlsx",
    "ke_hoach": TEMPLATES_DIR / "mau_import_ke_hoach.xlsx"
}

def load_data_from_db():
    """Đọc dữ liệu danh mục mới nhất từ SQLite CSDL"""
    print(f"Kết nối đến cơ sở dữ liệu: {DB_PATH}...")
    if not DB_PATH.exists():
        print(f"Lỗi: Không tìm thấy file database tại {DB_PATH}")
        sys.exit(1)
        
    conn = sqlite3.connect(str(DB_PATH))
    
    # 1. Đọc danh mục bưu cục (bao gồm cả mã 6 số, mã xã 4 số và mã cụm)
    # Lấy ma_bc và ten_buu_cuc để đối chiếu
    df_buucuc = pd.read_sql_query("""
        SELECT ma_bc as [Mã bưu cục], ten_buu_cuc as [Tên bưu cục] 
        FROM dim_buucuc 
        ORDER BY length(ma_bc), ma_bc
    """, conn)
    
    # 2. Đọc danh mục dịch vụ (toàn bộ dim_dichvu)
    df_dichvu = pd.read_sql_query("""
        SELECT ma_dich_vu as [Mã dịch vụ], ten_dich_vu as [Tên dịch vụ], nhom_dich_vu as [Nhóm dịch vụ]
        FROM dim_dichvu 
        ORDER BY nhom_chinh, ma_dich_vu
    """, conn)
    
    # 3. Đọc nhóm chính (BCCP, HCC, TCBC, PPBL, PHBC) cho phần kế hoạch
    df_nhomchinh = pd.read_sql_query("""
        SELECT DISTINCT nhom_chinh as [Nhóm chính]
        FROM dim_dichvu 
        WHERE nhom_chinh IS NOT NULL AND nhom_chinh != ''
        ORDER BY nhom_chinh
    """, conn)
    
    conn.close()
    
    print(f"Đọc thành công {len(df_buucuc)} bưu cục, {len(df_dichvu)} dịch vụ từ DB.")
    return df_buucuc, df_dichvu, df_nhomchinh

def update_template_references(filepath, sheet_name, df_data):
    """Cập nhật dữ liệu tham chiếu vào sheet chỉ định trong file template"""
    print(f"Đang ghi đè dữ liệu tham chiếu vào sheet '{sheet_name}' của tệp: {filepath.name}...")
    wb = openpyxl.load_workbook(filepath)
    
    # Nếu sheet chưa tồn tại, tạo mới
    if sheet_name in wb.sheetnames:
        # Xóa sheet cũ và tạo lại để đảm bảo sạch dữ liệu thừa
        std = wb[sheet_name]
        wb.remove(std)
        
    ws = wb.create_sheet(sheet_name)
    
    # Set định dạng chuỗi (@) cho cột đầu tiên để Excel không tự chuyển mã bưu cục thành số khi ghi
    ws.column_dimensions['A'].number_format = '@'
    
    # Ghi tiêu đề
    headers = list(df_data.columns)
    ws.append(headers)
    
    # Ghi dữ liệu
    for _, row in df_data.iterrows():
        # Chuyển đổi mã đầu tiên (mã bưu cục hoặc mã dịch vụ) sang chuỗi sạch
        vals = list(row.values)
        if vals[0] is not None:
            vals[0] = str(vals[0]).strip()
        ws.append(vals)
        
    # Cấu hình lại number_format dạng văn bản cho toàn bộ dữ liệu cột A trong sheet Ref
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=1).number_format = '@'
        
    wb.save(filepath)
    wb.close()
    print(f"Đã cập nhật xong sheet '{sheet_name}' ({ws.max_row} dòng).")

def apply_formulas_dich_vu_khac(filepath):
    """Nâng cấp toàn bộ công thức đối chiếu thông minh cho file mẫu dịch vụ khác"""
    print(f"Nâng cấp công thức đối chiếu thông minh cho tệp {filepath.name}...")
    wb = openpyxl.load_workbook(filepath)
    ws = wb['DuLieu']
    max_row = ws.max_row
    
    # Cập nhật công thức cho toàn bộ dòng từ 2 trở đi
    for row in range(2, max_row + 1):
        # Cột L (Kiểm tra bưu cục) - Sử dụng công thức so khớp thông minh xử lý cả Số/Chữ
        ws[f'L{row}'] = (
            f'=IF(B{row}="","",IF(OR(ISNUMBER(MATCH(B{row},Ref_BuuCuc!$A$2:$A$10000,0)),'
            f'ISNUMBER(MATCH(B{row}&"",Ref_BuuCuc!$A$2:$A$10000,0)),'
            f'IFERROR(ISNUMBER(MATCH(VALUE(B{row}),Ref_BuuCuc!$A$2:$A$10000,0)),FALSE)),"OK","Sai mã bưu cục"))'
        )
        # Cột M (Kiểm tra dịch vụ) - Tự động đối chiếu theo mã dịch vụ
        ws[f'M{row}'] = (
            f'=IF(C{row}="","",IF(OR(ISNUMBER(MATCH(C{row},Ref_DichVu!$A$2:$A$2000,0)),'
            f'ISNUMBER(MATCH(C{row}&"",Ref_DichVu!$A$2:$A$2000,0))),"OK","Sai mã dịch vụ"))'
        )
        
    wb.save(filepath)
    wb.close()
    print(f"Đã cập nhật công thức cho {max_row - 1} dòng dữ liệu của {filepath.name}.")

def apply_formulas_doanh_thu_bccp(filepath):
    """Nâng cấp toàn bộ công thức đối chiếu thông minh cho file mẫu doanh thu BCCP"""
    print(f"Nâng cấp công thức đối chiếu thông minh cho tệp {filepath.name}...")
    wb = openpyxl.load_workbook(filepath)
    ws = wb['DuLieu_DoanhThu']
    max_row = ws.max_row
    
    # Cập nhật công thức cho toàn bộ dòng từ 2 trở đi
    for row in range(2, max_row + 1):
        # Cột M (Kiểm tra bưu cục) - Sử dụng công thức so khớp thông minh xử lý cả Số/Chữ
        ws[f'M{row}'] = (
            f'=IF(D{row}="","",IF(OR(ISNUMBER(MATCH(D{row},Ref_BuuCuc!$A$2:$A$10000,0)),'
            f'ISNUMBER(MATCH(D{row}&"",Ref_BuuCuc!$A$2:$A$10000,0)),'
            f'IFERROR(ISNUMBER(MATCH(VALUE(D{row}),Ref_BuuCuc!$A$2:$A$10000,0)),FALSE)),"OK","Sai mã bưu cục"))'
        )
        # Cột N (Kiểm tra sản phẩm) - Tự động đối chiếu theo mã sản phẩm
        ws[f'N{row}'] = (
            f'=IF(E{row}="","",IF(OR(ISNUMBER(MATCH(E{row},Ref_DichVu!$A$2:$A$2000,0)),'
            f'ISNUMBER(MATCH(E{row}&"",Ref_DichVu!$A$2:$A$2000,0))),"OK","Sai mã sản phẩm"))'
        )
        
    wb.save(filepath)
    wb.close()
    print(f"Đã cập nhật công thức cho {max_row - 1} dòng dữ liệu của {filepath.name}.")

def apply_formulas_ke_hoach(filepath):
    """Nâng cấp toàn bộ công thức đối chiếu thông minh cho file mẫu kế hoạch"""
    print(f"Nâng cấp công thức đối chiếu thông minh cho tệp {filepath.name}...")
    wb = openpyxl.load_workbook(filepath)
    ws = wb['DuLieu_KeHoach']
    max_row = ws.max_row
    
    # Cập nhật công thức cho toàn bộ dòng từ 2 trở đi
    for row in range(2, max_row + 1):
        # Cột F (Kiểm tra đơn vị) - Sử dụng công thức so khớp thông minh xử lý cả Số/Chữ
        ws[f'F{row}'] = (
            f'=IF(D{row}="","",IF(OR(ISNUMBER(MATCH(D{row},Ref_BuuCuc!$A$2:$A$10000,0)),'
            f'ISNUMBER(MATCH(D{row}&"",Ref_BuuCuc!$A$2:$A$10000,0)),'
            f'IFERROR(ISNUMBER(MATCH(VALUE(D{row}),Ref_BuuCuc!$A$2:$A$10000,0)),FALSE)),"OK","Sai mã đơn vị"))'
        )
        # Cột G (Kiểm tra nhóm) - Tự động đối chiếu theo mã nhóm chính
        ws[f'G{row}'] = (
            f'=IF(B{row}="","",IF(OR(ISNUMBER(MATCH(B{row},Ref_NhomChinh!$A$2:$A$100,0)),'
            f'ISNUMBER(MATCH(B{row}&"",Ref_NhomChinh!$A$2:$A$100,0))),"OK","Sai nhóm chính"))'
        )
        
    wb.save(filepath)
    wb.close()
    print(f"Đã cập nhật công thức cho {max_row - 1} dòng dữ liệu của {filepath.name}.")

def main():
    print("=== BẮT ĐẦU ĐỒNG BỘ MẪU FILE EXCEL VÀ NÂNG CẤP CÔNG THỨC ===")
    
    # 1. Đọc dữ liệu từ DB
    df_buucuc, df_dichvu, df_nhomchinh = load_data_from_db()
    
    # 2. Cập nhật sheet Ref cho mau_import_dich_vu_khac.xlsx
    f_dich_vu_khac = TEMPLATE_FILES["dich_vu_khac"]
    if f_dich_vu_khac.exists():
        update_template_references(f_dich_vu_khac, "Ref_BuuCuc", df_buucuc)
        update_template_references(f_dich_vu_khac, "Ref_DichVu", df_dichvu)
        apply_formulas_dich_vu_khac(f_dich_vu_khac)
    else:
        print(f"Cảnh báo: Không tìm thấy file mẫu tại {f_dich_vu_khac}")

    # 3. Cập nhật sheet Ref cho mau_import_doanh_thu_BCCP.xlsx
    f_bccp = TEMPLATE_FILES["doanh_thu_bccp"]
    if f_bccp.exists():
        update_template_references(f_bccp, "Ref_BuuCuc", df_buucuc)
        update_template_references(f_bccp, "Ref_DichVu", df_dichvu)
        apply_formulas_doanh_thu_bccp(f_bccp)
    else:
        print(f"Cảnh báo: Không tìm thấy file mẫu tại {f_bccp}")

    # 4. Cập nhật sheet Ref cho mau_import_ke_hoach.xlsx
    f_ke_hoach = TEMPLATE_FILES["ke_hoach"]
    if f_ke_hoach.exists():
        update_template_references(f_ke_hoach, "Ref_BuuCuc", df_buucuc)
        update_template_references(f_ke_hoach, "Ref_NhomChinh", df_nhomchinh)
        apply_formulas_ke_hoach(f_ke_hoach)
    else:
        print(f"Cảnh báo: Không tìm thấy file mẫu tại {f_ke_hoach}")

    print("=== HOÀN THÀNH ĐỒNG BỘ MẪU FILE EXCEL & CẬP NHẬT CÔNG THỨC THÀNH CÔNG ===")

if __name__ == "__main__":
    main()
