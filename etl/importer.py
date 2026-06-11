# -*- coding: utf-8 -*-
"""
Module ETL Importer - Import dữ liệu từ file Excel (.xlsx) vào SQLite.
"""

import os
import sys
import re
import sqlite3
import logging
from datetime import datetime
import openpyxl
from config.settings import (
    COLUMN_NAMES, ABNORMAL_CMS, EXCEL_DATA_START_ROW, EXCEL_NUM_COLS
)

logger = logging.getLogger(__name__)

# BATCH SIZE cho INSERT
BATCH_SIZE = 5000

def _parse_date(date_str):
    """
    Chuyển đổi ngày từ định dạng dd/mm/yyyy → YYYY-MM-DD (ISO).
    Hỗ trợ cả datetime, pandas.Timestamp, và chuỗi.
    """
    if date_str is None:
        return None
    
    # pandas.Timestamp kế thừa từ datetime    
    if isinstance(date_str, datetime):
        return date_str.strftime('%Y-%m-%d')
    
    # Xử lý pandas Timestamp nếu không bắt được ở trên
    try:
        import pandas as pd
        if isinstance(date_str, pd.Timestamp):
            return date_str.strftime('%Y-%m-%d')
    except ImportError:
        pass
        
    date_str = str(date_str).strip()
    if not date_str:
        return None
        
    # Thử parse dd/mm/yyyy
    try:
        dt = datetime.strptime(date_str, '%d/%m/%Y')
        return dt.strftime('%Y-%m-%d')
    except ValueError:
        pass
        
    # Thử parse yyyy-mm-dd (có thể kèm giờ phút giây)
    try:
        dt = datetime.strptime(date_str[:10], '%Y-%m-%d')
        return dt.strftime('%Y-%m-%d')
    except ValueError:
        pass
        
    logger.warning(f"Không parse được ngày: '{date_str}'")
    return None

def _detect_month(filename):
    """
    Tự động phát hiện tháng từ tên file Excel (ví dụ T4 -> T04).
    """
    match = re.match(r'^T(\d{1,2})\s*[-_\s]', filename, re.IGNORECASE)
    if match:
        month_num = int(match.group(1))
        if 1 <= month_num <= 12:
            return f'T{month_num:02d}'
            
    # Thử tìm chữ T + số bất kỳ trong tên file
    match_any = re.search(r'[T](\d{1,2})', filename, re.IGNORECASE)
    if match_any:
        month_num = int(match_any.group(1))
        if 1 <= month_num <= 12:
            return f'T{month_num:02d}'
            
    return None

def _safe_float(value):
    if value is None:
        return 0.0
    try:
        return float(value)
    except (ValueError, TypeError):
        return 0.0

def _safe_int(value):
    if value is None:
        return 0
    try:
        return int(value)
    except (ValueError, TypeError):
        return 0


def _get_db_connection(db_path):
    conn = sqlite3.connect(db_path)
    # Da tat che do vua doc vua ghi (WAL) theo yeu cau cua Sep de tranh loi lock tren OneDrive
    conn.execute("PRAGMA journal_mode=delete;")
    conn.execute("PRAGMA busy_timeout=30000;")
    return conn

def _auto_aggregate_after_import(db_path, service_type, years_months_pairs):
    """
    Tự động gộp lại số liệu tháng/tuần ngay sau khi nạp thành công.
    """
    import logging
    logger = logging.getLogger(__name__)
    try:
        from etl.aggregator import rebuild_monthly, rebuild_monthly_customer, rebuild_weekly, rebuild_plans_weekly
        conn = _get_db_connection(db_path)
        try:
            for nam, thang_int in years_months_pairs:
                logger.info(f"[Auto-Refresh] Đang gộp lại số liệu cho {service_type} - T{thang_int:02d}/{nam}...")
                
                # 1. Gộp doanh thu tháng
                rebuild_monthly(conn, nam, thang_int)
                if service_type in ('BCCP', 'RAW'):
                    rebuild_monthly_customer(conn, nam, thang_int)
                    
                # 2. Gộp doanh thu tuần (weekly)
                rebuild_weekly(conn, nam)
                
                # 3. Gộp kế hoạch tuần nếu là nạp Kế hoạch
                if service_type == 'PLAN':
                    rebuild_plans_weekly(conn, nam)
                    
                logger.info(f"[Auto-Refresh] Đã hoàn tất gộp số liệu cho T{thang_int:02d}/{nam}")
        finally:
            conn.close()
    except Exception as e:
        logger.error(f"[Auto-Refresh] Lỗi khi tự động gộp số liệu: {e}", exc_info=True)

def import_excel_file(db_path, excel_path, import_batch=None, thang=None, mode='append'):
    """
    Import 1 file Excel (.xlsx) vào SQLite.
    """
    filename = os.path.basename(excel_path)
    logger.info(f"Bắt đầu import file: {filename}")
    
    # Phát hiện tháng
    if thang is None:
        thang = _detect_month(filename)
        if thang is None:
            # Thử thư mục cha
            parent_dir = os.path.basename(os.path.dirname(excel_path))
            thang = _detect_month(parent_dir)
            
    if thang is None:
        thang = "T01"  # Mặc định fallback
        logger.warning(f"Không phát hiện được tháng cho file: {filename}, fallback sang T01")
        
    # Tạo batch_id
    batch_id = import_batch or (datetime.now().strftime('%Y%m%d_%H%M%S') + '_' + filename)
    

    # Nếu là overwrite, đọc qua file để lấy danh mục sản phẩm và tháng/năm xóa trước
    if mode == 'overwrite':
        logger.info("Chế độ ghi đè: Đang xác định các sản phẩm và tháng/năm để xóa dữ liệu cũ...")
        try:
            wb_temp = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
            ws_temp = wb_temp.active
            to_delete = set()
            for r_idx, row in enumerate(ws_temp.iter_rows(min_row=EXCEL_DATA_START_ROW, max_col=10, values_only=True), start=EXCEL_DATA_START_ROW):
                if row[0] is None:
                    continue
                buu_cuc = str(row[3]).strip() if row[3] is not None else None
                sp_dv = str(row[4]).strip() if row[4] is not None else None
                ngay_cn = _parse_date(row[5])
                if sp_dv and ngay_cn and buu_cuc:
                    y = int(ngay_cn[:4])
                    m_str = f"T{int(ngay_cn[5:7]):02d}"
                    to_delete.add((sp_dv, buu_cuc, m_str, y))
            wb_temp.close()
            
            if to_delete:
                conn_del = _get_db_connection(db_path)
                cursor_del = conn_del.cursor()
                for sp, bc, m_str, y in to_delete:
                    logger.info(f"Xóa dữ liệu cũ sản phẩm '{sp}' bưu cục '{bc}' - {m_str}/{y} trong transactions...")
                    cursor_del.execute("""
                        DELETE FROM transactions 
                        WHERE san_pham_dv = ? AND buu_cuc = ? AND thang_du_lieu = ? AND nam_du_lieu = ?
                    """, (sp, bc, m_str, y))
                conn_del.commit()
                conn_del.close()
        except Exception as ex_del:
            logger.error(f"Lỗi khi xóa dữ liệu cũ BCCP: {ex_del}")

    # Mở file Excel
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active
    
    conn = _get_db_connection(db_path)
    cursor = conn.cursor()
    
    insert_cols = [
        'thang_du_lieu', 'nam_du_lieu', 'import_batch', 'stt', 'cms', 'ma_hop_dong', 'buu_cuc',
        'san_pham_dv', 'ngay_chap_nhan', 'san_luong',
        'khoi_luong_thuc', 'khoi_luong_tinh_cuoc',
        'cuoc_cb_cp', 'cuoc_cb_gtgt', 'cuoc_cb_cod', 'cuoc_cb_tong',
        'cuoc_tt_cp', 'cuoc_tt_gtgt', 'cuoc_tt_cod', 'cuoc_tt_tong',
        'thue_vat', 'cuoc_tt_gom_vat', 'cuoc_chenh_lech',
        'tien_cod', 'nho_thu_khac'
    ]
    placeholders = ', '.join(['?'] * len(insert_cols))
    col_names_sql = ', '.join(insert_cols)
    insert_sql = f"INSERT OR IGNORE INTO transactions ({col_names_sql}) VALUES ({placeholders})"
    
    total_rows = 0
    inserted = 0
    warnings = []
    batch_buffer = []
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=EXCEL_DATA_START_ROW, 
                                                 max_col=ws.max_column,
                                                 values_only=True), 
                                    start=EXCEL_DATA_START_ROW):
        # Bỏ qua dòng trống
        if row[0] is None:
            continue
            
        total_rows += 1
        
        stt = _safe_int(row[0])
        cms = str(row[1]).strip() if row[1] is not None and str(row[1]).strip() != '' else None
        ma_hop_dong = str(row[2]).strip() if row[2] is not None and str(row[2]).strip() != '' else None
        buu_cuc = str(row[3]).strip() if row[3] is not None else None
        san_pham_dv = str(row[4]).strip() if row[4] is not None else None
        ngay_chap_nhan = _parse_date(row[5])
        san_luong = _safe_int(row[6])
        khoi_luong_thuc = _safe_float(row[7]) if len(row) > 7 else 0.0
        khoi_luong_tinh_cuoc = _safe_float(row[8]) if len(row) > 8 else 0.0
        
        # Hỗ trợ cả 2 chuẩn template: Mới (12 cột) và Cũ (22 cột)
        if ws.max_column <= 12:
            cuoc_tt_tong = _safe_float(row[9]) if len(row) > 9 else 0.0
            thue_vat = _safe_float(row[10]) if len(row) > 10 else 0.0
            cuoc_tt_gom_vat = _safe_float(row[11]) if len(row) > 11 else 0.0
            
            cuoc_cb_cp = cuoc_cb_gtgt = cuoc_cb_cod = cuoc_cb_tong = 0.0
            cuoc_tt_cp = cuoc_tt_gtgt = cuoc_tt_cod = 0.0
            cuoc_chenh_lech = tien_cod = nho_thu_khac = 0.0
        else:
            cuoc_cb_cp = _safe_float(row[9]) if len(row) > 9 else 0.0
            cuoc_cb_gtgt = _safe_float(row[10]) if len(row) > 10 else 0.0
            cuoc_cb_cod = _safe_float(row[11]) if len(row) > 11 else 0.0
            cuoc_cb_tong = _safe_float(row[12]) if len(row) > 12 else 0.0
            cuoc_tt_cp = _safe_float(row[13]) if len(row) > 13 else 0.0
            cuoc_tt_gtgt = _safe_float(row[14]) if len(row) > 14 else 0.0
            cuoc_tt_cod = _safe_float(row[15]) if len(row) > 15 else 0.0
            cuoc_tt_tong = _safe_float(row[16]) if len(row) > 16 else 0.0
            thue_vat = _safe_float(row[17]) if len(row) > 17 else 0.0
            cuoc_tt_gom_vat = _safe_float(row[18]) if len(row) > 18 else 0.0
            cuoc_chenh_lech = _safe_float(row[19]) if len(row) > 19 else 0.0
            tien_cod = _safe_float(row[20]) if len(row) > 20 else 0.0
            nho_thu_khac = _safe_float(row[21]) if len(row) > 21 else 0.0
        
        # Xử lý CMS vãng lai
        if cms is None:
            if buu_cuc:
                cms = f'VANGLAI_{buu_cuc}'
            else:
                cms = 'VANGLAI_UNKNOWN'
                warnings.append(f"Dòng {row_idx}: CMS null và bưu cục cũng null")
                
        # Cảnh báo CMS bất thường
        if cms in ABNORMAL_CMS:
            warnings.append(f"Dòng {row_idx}: CMS bất thường (tên người): '{cms}'")
            
        # Trích năm từ ngày chấp nhận
        nam = int(ngay_chap_nhan[:4]) if ngay_chap_nhan and len(ngay_chap_nhan) >= 4 else None
        
        row_data = (
            thang, nam, batch_id, stt, cms, ma_hop_dong, buu_cuc,
            san_pham_dv, ngay_chap_nhan, san_luong,
            khoi_luong_thuc, khoi_luong_tinh_cuoc,
            cuoc_cb_cp, cuoc_cb_gtgt, cuoc_cb_cod, cuoc_cb_tong,
            cuoc_tt_cp, cuoc_tt_gtgt, cuoc_tt_cod, cuoc_tt_tong,
            thue_vat, cuoc_tt_gom_vat, cuoc_chenh_lech,
            tien_cod, nho_thu_khac
        )
        batch_buffer.append(row_data)
        
        if len(batch_buffer) >= BATCH_SIZE:
            cursor.executemany(insert_sql, batch_buffer)
            inserted += cursor.rowcount
            conn.commit()
            batch_buffer = []
            
    if batch_buffer:
        cursor.executemany(insert_sql, batch_buffer)
        inserted += cursor.rowcount
        conn.commit()
        
    skipped = total_rows - inserted
    wb.close()
    
    # Kiểm tra thiếu mapping sản phẩm/bưu cục
    missing = check_missing_mappings(conn)
    if missing['missing_products']:
        warnings.append(f"Mã sản phẩm mới chưa phân nhóm: {', '.join(missing['missing_products'][:3])}")
    if missing['missing_post_offices']:
        warnings.append(f"Mã bưu cục mới chưa phân cụm: {', '.join(missing['missing_post_offices'][:3])}")
        
    conn.close()
    
    # Auto-refresh summary tables cho tháng vừa import
    try:
        thang_int = int(thang[1:]) if thang.startswith('T') else int(thang)
        _auto_aggregate_after_import(db_path, 'BCCP', [(nam, thang_int)])
    except Exception as ex_ref:
        logger.error(f"Lỗi tự động refresh sau import Template: {ex_ref}")
        
    return {
        'batch_id': batch_id,
        'file': filename,
        'thang': thang,
        'total_rows': total_rows,
        'inserted': inserted,
        'skipped': skipped,
        'warnings': warnings
    }

def clean_str_field(val):
    import pandas as pd
    if pd.isna(val):
        return ""
    val_str = str(val).strip()
    if val_str.endswith(".0"):
        try:
            float(val_str)
            val_str = val_str[:-2]
        except ValueError:
            pass
    return val_str

def import_raw_excel_file(db_path, excel_path, import_batch=None, thang=None, mode='append'):
    """
    Import file Excel dữ liệu chi tiết bưu gửi (RAW) từ CAS.
    Sẽ thực hiện đọc, trích xuất dòng, nén dữ liệu (group by) và chèn vào SQLite.
    Tháng dữ liệu được tự động xác định từ ngày chấp nhận của dòng chi tiết đầu tiên.
    Sản lượng = số mã bưu gửi duy nhất (nunique) trong mỗi nhóm.
    """
    import pandas as pd
    import gc
    filename = os.path.basename(excel_path)
    logger.info(f"Bắt đầu import RAW file: {filename}")
        
    batch_id = import_batch or (datetime.now().strftime('%Y%m%d_%H%M%S') + '_RAW_' + filename)
    
    # 1. Đọc từng sheet để thu thập all_records
    all_records = []
    total_ignored = 0
    warnings = []
    first_date_str = None  # Để phát hiện tháng từ dữ liệu
    
    try:
        xl = pd.ExcelFile(excel_path)
        sheets = xl.sheet_names
        
        current_cms = None
        for sheetname in sheets:
            engine = "openpyxl" if excel_path.endswith(".xlsx") else None
            df = pd.read_excel(excel_path, sheet_name=sheetname, header=None, engine=engine)
            
            start_row = 0
            for idx in range(min(50, len(df))):
                val_a = str(df.iloc[idx, 0]).strip() if pd.notna(df.iloc[idx, 0]) else ""
                if "1. Chi tiết sản lượng phát sinh" in val_a:
                    start_row = idx
                    break
                    
            for idx in range(start_row, len(df)):
                row_vals = df.iloc[idx].tolist()
                if len(row_vals) < 32:
                    total_ignored += 1
                    continue
                    
                val_a = str(row_vals[0]).strip() if pd.notna(row_vals[0]) else ""
                val_b = clean_str_field(row_vals[1])
                val_d = str(row_vals[3]).strip() if pd.notna(row_vals[3]) else ""
                
                # Bỏ qua dòng tiêu đề section
                if "1. Chi tiết sản lượng phát sinh" in val_a or "2. Chi tiết sản lượng điều chỉnh" in val_a:
                    continue
                    
                # Nhận diện dòng group CMS
                if val_a != "" and val_d == "":
                    current_cms = val_b
                    continue
                    
                # Dòng chi tiết bưu gửi
                is_valid_bg = val_d != "" and val_d != "Số hiệu bưu gửi" and not (val_d.startswith("(") and val_d.endswith(")"))
                if is_valid_bg:
                    if current_cms is None:
                        total_ignored += 1
                        continue
                        
                    def get_num(val):
                        if pd.isna(val): return 0.0
                        try: return float(val)
                        except: return 0.0
                    
                    raw_date_val = row_vals[4] if pd.notna(row_vals[4]) else None
                    parsed_date = _parse_date(raw_date_val)
                    
                    # Ghi nhận ngày đầu tiên để xác định tháng dữ liệu
                    if first_date_str is None and parsed_date:
                        first_date_str = parsed_date
                        
                    record = {
                        'CMS': current_cms,
                        'Ma_HD': clean_str_field(row_vals[7]),
                        'Buu_Cuc': clean_str_field(row_vals[5]),
                        'San_Pham': clean_str_field(row_vals[9]),
                        'Ngay_CN': parsed_date,
                        'Ma_BG': val_d,
                        'KL_Thuc': get_num(row_vals[13]),
                        'KL_TinhCuoc': get_num(row_vals[14]),
                        'CB_CP': get_num(row_vals[15]),
                        'CB_GTGT': get_num(row_vals[16]),
                        'CB_COD': get_num(row_vals[17]),
                        'CB_Tong': get_num(row_vals[18]),
                        'TT_CP': get_num(row_vals[19]),
                        'TT_GTGT': get_num(row_vals[20]),
                        'TT_COD': get_num(row_vals[21]),
                        'TT_Tong': get_num(row_vals[22]),
                        'TT_VAT': get_num(row_vals[23]),
                        'TT_TongVAT': get_num(row_vals[24]),
                        'Cuoc_CL': get_num(row_vals[25]),
                        'COD_Tien': get_num(row_vals[30]),
                        'COD_Khac': get_num(row_vals[31])
                    }
                    all_records.append(record)
                else:
                    total_ignored += 1
                    
            del df
            gc.collect()
    except Exception as e:
        logger.error(f"Lỗi xử lý file RAW {filename}: {e}")
        return {'batch_id': batch_id, 'file': filename, 'thang': 'N/A', 'total_rows': 0, 'inserted': 0, 'skipped': 0, 'warnings': [str(e)]}

    if not all_records:
        return {'batch_id': batch_id, 'file': filename, 'thang': 'N/A', 'total_rows': 0, 'inserted': 0, 'skipped': 0, 'warnings': ['Không tìm thấy dữ liệu phát sinh']}

    # Xác định tháng từ ngày chấp nhận dòng đầu tiên (VD: 2026-01-06 → T01)
    if thang is None:
        if first_date_str:
            try:
                month_num = int(first_date_str.split('-')[1])
                thang = f'T{month_num:02d}'
                logger.info(f"Phát hiện tháng từ dữ liệu: {thang} (ngày đầu: {first_date_str})")
            except (IndexError, ValueError):
                thang = "T01"
                warnings.append(f"Không parse được tháng từ ngày '{first_date_str}', fallback T01")
        else:
            thang = "T01"
            warnings.append("Không tìm thấy ngày chấp nhận trong dữ liệu, fallback T01")


    # Nếu là overwrite, xóa dữ liệu cũ trùng khớp dịch vụ và khoảng ngày trong file
    if mode == 'overwrite' and len(all_records) > 0:
        logger.info("Chế độ ghi đè RAW: Đang xác định các sản phẩm và tháng/năm để xóa dữ liệu cũ...")
        try:
            to_delete = set()
            for rec in all_records:
                sp = rec['San_Pham']
                ngay_cn = rec['Ngay_CN']
                buu_cuc = rec['Buu_Cuc']
                if sp and ngay_cn and buu_cuc:
                    y = int(ngay_cn[:4])
                    m_str = f"T{int(ngay_cn[5:7]):02d}"
                    to_delete.add((sp, buu_cuc, m_str, y))
            
            if to_delete:
                conn_del = _get_db_connection(db_path)
                cursor_del = conn_del.cursor()
                for sp, m_str, y in to_delete:
                    logger.info(f"Xóa dữ liệu cũ RAW của sản phẩm '{sp}' - {m_str}/{y} trong transactions...")
                    cursor_del.execute("""
                        DELETE FROM transactions 
                        WHERE san_pham_dv = ? AND thang_du_lieu = ? AND nam_du_lieu = ?
                    """, (sp, m_str, y))
                conn_del.commit()
                conn_del.close()
        except Exception as ex_del:
            logger.error(f"Lỗi khi xóa dữ liệu cũ RAW: {ex_del}")

    # 2. Gom nhóm — san_luong = nunique(Ma_BG) đảm bảo chỉ đếm số hiệu duy nhất
    df_all = pd.DataFrame(all_records)
    groupby_cols = ['CMS', 'Ma_HD', 'Buu_Cuc', 'San_Pham', 'Ngay_CN']
    agg_dict = {
        'Ma_BG': 'nunique',
        'KL_Thuc': 'sum', 'KL_TinhCuoc': 'sum',
        'CB_CP': 'sum', 'CB_GTGT': 'sum', 'CB_COD': 'sum', 'CB_Tong': 'sum',
        'TT_CP': 'sum', 'TT_GTGT': 'sum', 'TT_COD': 'sum', 'TT_Tong': 'sum',
        'TT_VAT': 'sum', 'TT_TongVAT': 'sum', 'Cuoc_CL': 'sum',
        'COD_Tien': 'sum', 'COD_Khac': 'sum'
    }
    df_grouped = df_all.groupby(groupby_cols, as_index=False).agg(agg_dict)
    
    # 3. Ghi vào SQLite
    conn = _get_db_connection(db_path)
    cursor = conn.cursor()
    
    # Trích năm từ tháng (VD: T01 → lấy từ first_date_str)
    nam_du_lieu = int(first_date_str[:4]) if first_date_str and len(first_date_str) >= 4 else None
    
    insert_cols = [
        'thang_du_lieu', 'nam_du_lieu', 'import_batch', 'stt', 'cms', 'ma_hop_dong', 'buu_cuc',
        'san_pham_dv', 'ngay_chap_nhan', 'san_luong',
        'khoi_luong_thuc', 'khoi_luong_tinh_cuoc',
        'cuoc_cb_cp', 'cuoc_cb_gtgt', 'cuoc_cb_cod', 'cuoc_cb_tong',
        'cuoc_tt_cp', 'cuoc_tt_gtgt', 'cuoc_tt_cod', 'cuoc_tt_tong',
        'thue_vat', 'cuoc_tt_gom_vat', 'cuoc_chenh_lech',
        'tien_cod', 'nho_thu_khac'
    ]
    placeholders = ', '.join(['?'] * len(insert_cols))
    insert_sql = f"INSERT OR IGNORE INTO transactions ({', '.join(insert_cols)}) VALUES ({placeholders})"
    
    batch_buffer = []
    inserted = 0
    
    for idx, row in df_grouped.iterrows():
        cms_val = row['CMS'] if pd.notna(row['CMS']) and str(row['CMS']).strip() != '' else f"VANGLAI_{row['Buu_Cuc']}"
        # Chuyển "" → None để khớp với NULL trong DB (SQLite: NULL ≠ "" → sẽ gây duplicate nếu không sửa)
        ma_hd_val = row['Ma_HD'] if row['Ma_HD'] != '' else None

        row_data = (
            thang, nam_du_lieu, batch_id, idx + 1, cms_val, ma_hd_val, row['Buu_Cuc'],
            row['San_Pham'], row['Ngay_CN'], row['Ma_BG'],
            row['KL_Thuc'], row['KL_TinhCuoc'],
            row['CB_CP'], row['CB_GTGT'], row['CB_COD'], row['CB_Tong'],
            row['TT_CP'], row['TT_GTGT'], row['TT_COD'], row['TT_Tong'],
            row['TT_VAT'], row['TT_TongVAT'], row['Cuoc_CL'],
            row['COD_Tien'], row['COD_Khac']
        )
        batch_buffer.append(row_data)
        
        if len(batch_buffer) >= BATCH_SIZE:
            cursor.executemany(insert_sql, batch_buffer)
            inserted += cursor.rowcount
            conn.commit()
            batch_buffer = []
            
    if batch_buffer:
        cursor.executemany(insert_sql, batch_buffer)
        inserted += cursor.rowcount
        conn.commit()
        
    # Kiểm tra thiếu mapping sản phẩm/bưu cục
    missing = check_missing_mappings(conn)
    if missing['missing_products']:
        warnings.append(f"Mã sản phẩm mới chưa phân nhóm: {', '.join(missing['missing_products'][:3])}")
    if missing['missing_post_offices']:
        warnings.append(f"Mã bưu cục mới chưa phân cụm: {', '.join(missing['missing_post_offices'][:3])}")

    conn.close()
    
    skipped = len(df_grouped) - inserted
    
    # Auto-refresh summary tables cho tháng vừa import
    try:
        thang_int = int(thang[1:]) if thang.startswith('T') else int(thang)
        _auto_aggregate_after_import(db_path, 'BCCP', [(nam_du_lieu, thang_int)])
    except Exception as ex_ref:
        logger.error(f"Lỗi tự động refresh sau import RAW: {ex_ref}")
        
    return {
        'batch_id': batch_id,
        'file': filename,
        'thang': thang,
        'total_rows': len(df_grouped),
        'inserted': inserted,
        'skipped': skipped,
        'warnings': warnings
    }

def check_missing_mappings(db_path_or_conn):
    """
    Kiểm tra xem có mã sản phẩm hoặc bưu cục nào trong transactions chưa được định nghĩa trong bảng dim.
    Trả về dict chứa danh sách các mã thiếu.
    """
    if isinstance(db_path_or_conn, sqlite3.Connection):
        conn = db_path_or_conn
        should_close = False
    else:
        conn = _get_db_connection(db_path_or_conn)
        should_close = True
        
    cursor = conn.cursor()
    
    missing_sp = []
    missing_bc = []
    
    try:
        # Kiểm tra sản phẩm
        cursor.execute("""
            SELECT DISTINCT t.san_pham_dv 
            FROM transactions t 
            LEFT JOIN dim_spdv d ON t.san_pham_dv = d.ma_spdv 
            WHERE d.ma_spdv IS NULL AND t.san_pham_dv IS NOT NULL AND t.san_pham_dv != ''
        """)
        missing_sp = [r[0] for r in cursor.fetchall()]
        
        # Kiểm tra bưu cục
        cursor.execute("""
            SELECT DISTINCT t.buu_cuc 
            FROM transactions t 
            LEFT JOIN dim_buucuc b ON t.buu_cuc = b.ma_bc 
            WHERE b.ma_bc IS NULL AND t.buu_cuc IS NOT NULL AND t.buu_cuc != ''
        """)
        missing_bc = [r[0] for r in cursor.fetchall()]
    except sqlite3.Error as e:
        logger.error(f"Lỗi khi kiểm tra danh mục chưa phân loại: {e}")
    finally:
        if should_close:
            conn.close()
        
    return {
        'missing_products': missing_sp,
        'missing_post_offices': missing_bc
    }

def import_any_excel_file(db_path, excel_path, import_batch=None, thang=None, mode='append'):
    """
    Hàm điều phối thông minh: Tự động phân tích cấu trúc file Excel để nhận diện
    đây là file dữ liệu thô (RAW từ CAS) hay file mẫu nén/thủ công (Template).
    Sau đó gọi hàm import tương ứng.
    """
    import pandas as pd
    filename = os.path.basename(excel_path)
    logger.info(f"Phân tích loại file để import: {filename}")
    
    # Xác định engine dựa trên định dạng file
    engine = "openpyxl" if excel_path.endswith(".xlsx") else "xlrd"
    
    try:
        # Đọc sheet đầu tiên (chỉ lấy 50 dòng đầu để phân tích cấu trúc rất nhanh)
        xl = pd.ExcelFile(excel_path, engine=engine)
        sheet_name = xl.sheet_names[0]
        df_sample = pd.read_excel(excel_path, sheet_name=sheet_name, nrows=50, header=None, engine=engine)
        
        # Tìm xem có chứa chuỗi chỉ thị của file RAW CAS không
        is_raw = False
        for idx in range(len(df_sample)):
            val_a = str(df_sample.iloc[idx, 0]).strip() if pd.notna(df_sample.iloc[idx, 0]) else ""
            if "1. Chi tiết sản lượng phát sinh" in val_a:
                is_raw = True
                break
                
        if is_raw:
            logger.info(f"Nhận diện file '{filename}' là dữ liệu thô (RAW) từ CAS. Tiến hành nén và import...")
            return import_raw_excel_file(db_path, excel_path, import_batch=import_batch, thang=thang, mode=mode)
        else:
            logger.info(f"Nhận diện file '{filename}' là tệp tin mẫu (Template). Tiến hành import trực tiếp...")
            # Nếu file template là .xls thì openpyxl hiện tại sẽ bị lỗi.
            # Ta có thể báo lỗi thân thiện nếu file template là .xls
            if excel_path.endswith(".xls"):
                raise ValueError("Tệp mẫu điền tay (Template) phải là định dạng .xlsx (Excel mới). Vui lòng lưu tệp mẫu dưới dạng .xlsx và thử lại.")
            return import_excel_file(db_path, excel_path, import_batch=import_batch, thang=thang, mode=mode)
            
    except Exception as e:
        logger.error(f"Lỗi phân loại file Excel {filename}: {e}")
        raise

def check_and_migrate_services_tables(conn):
    """
    Tự động kiểm tra và nâng cấp cấu trúc các bảng transactions_hcc/tcbc/ppbl/phbc
    để đồng bộ các cột khoảng ngày, ten_dich_vu, san_luong, stt.
    """
    cursor = conn.cursor()
    tables = ['transactions_hcc', 'transactions_tcbc', 'transactions_ppbl', 'transactions_phbc']
    needed_columns = {
        'ten_dich_vu': 'TEXT',
        'san_luong': 'INTEGER',
        'tu_ngay': 'INTEGER',
        'tu_thang': 'INTEGER',
        'tu_nam': 'INTEGER',
        'den_ngay': 'INTEGER',
        'den_thang': 'INTEGER',
        'den_nam': 'INTEGER',
        'stt': 'INTEGER'
    }
    
    for table in tables:
        # Lấy thông tin các cột hiện có
        columns = [row[1] for row in cursor.execute(f"PRAGMA table_info({table})").fetchall()]
        for col_name, col_type in needed_columns.items():
            if col_name not in columns:
                logger.info(f"Đang tự động nâng cấp bảng {table}: Thêm cột {col_name} ({col_type})...")
                cursor.execute(f"ALTER TABLE {table} ADD COLUMN {col_name} {col_type}")
    conn.commit()

def import_service_excel(db_path, excel_path, service_type, import_batch=None, thang=None, mode='append'):
    """
    Import file Excel dữ liệu dịch vụ khác (SERVICES) bao gồm HCC, TCBC, PPBL, PHBC.
    Dữ liệu được phân rã thành từng dòng ngày cụ thể (tu_ngay == den_ngay == ngày phân rã)
    trước khi ghi vào 4 bảng thô tương ứng (transactions_hcc/tcbc/ppbl/phbc).
    """
    import pandas as pd
    from datetime import date, timedelta
    import calendar
    
    filename = os.path.basename(excel_path)
    logger.info(f"Bắt đầu import dịch vụ {service_type}: {filename} (mode={mode})")
    
    # 1. Tự động phát hiện tháng nếu chưa truyền vào
    if thang is None:
        thang = _detect_month(filename)
        if thang is None:
            # Thử thư mục cha
            parent_dir = os.path.basename(os.path.dirname(excel_path))
            thang = _detect_month(parent_dir)
            
    if thang is None:
        thang = "T01"  # Fallback mặc định
        logger.warning(f"Không phát hiện được tháng cho file: {filename}, fallback sang T01")
        
    # Tạo batch_id thống nhất
    batch_id = import_batch or (datetime.now().strftime('%Y%m%d_%H%M%S') + f'_{service_type}_' + filename)
    nam_du_lieu = datetime.now().year
    
    try:
        # 2. Khởi tạo/Nâng cấp cấu trúc các bảng thô
        conn_init = _get_db_connection(db_path)
        check_and_migrate_services_tables(conn_init)
        
        # 3. Đọc mapping sản phẩm dịch vụ từ dim_dichvu
        cursor_init = conn_init.cursor()
        dim_dichvu = {}
        try:
            cursor_init.execute("SELECT ma_dich_vu, ten_dich_vu, nhom_chinh FROM dim_dichvu")
            for ma_dv, ten_dv, nhom_ch in cursor_init.fetchall():
                if ma_dv:
                    dim_dichvu[ma_dv.strip().upper()] = (ma_dv.strip(), ten_dv.strip(), nhom_ch.strip())
                if ten_dv:
                    dim_dichvu[ten_dv.strip().upper()] = (ma_dv.strip(), ten_dv.strip(), nhom_ch.strip())
        except sqlite3.Error as e:
            logger.error(f"Lỗi khi đọc bảng dim_dichvu: {e}")
        finally:
            conn_init.close()
            
        # 4. Đọc file Excel
        engine = "openpyxl" if str(excel_path).endswith(".xlsx") else "xlrd"
        df = pd.read_excel(excel_path, engine=engine)
        
        if len(df.columns) < 5:
            return {
                'batch_id': batch_id,
                'file': filename,
                'thang': thang,
                'total_rows': 0,
                'inserted': 0,
                'skipped': 0,
                'warnings': [f"File không đủ 5 cột cơ bản (STT, Mã BC, Dịch vụ, SL, DT). Số cột: {len(df.columns)}"]
            }
            
        total_rows = 0
        warnings = []
        records_by_table = {}  # { 'transactions_hcc': [row_data, ...] }
        
        # 5. Duyệt qua từng dòng và thực hiện phân rã ngày
        for idx, row in df.iterrows():
            stt_val = row.iloc[0]
            if pd.isna(stt_val) and idx > 0:
                continue  # Bỏ qua dòng trống
                
            if str(stt_val).strip().upper() == 'STT':
                continue
                
            total_rows += 1
            ma_bc = clean_str_field(row.iloc[1])
            ten_dv = clean_str_field(row.iloc[2])
            sl = _safe_int(row.iloc[3])
            dt = _safe_float(row.iloc[4])
            
            if not ma_bc or not ten_dv:
                warnings.append(f"Dòng {idx+2}: Thiếu mã bưu cục hoặc tên dịch vụ. Bỏ qua.")
                continue
                
            # Tra cứu thông tin dịch vụ trong dim_dichvu
            mapping = dim_dichvu.get(ten_dv.upper())
            if mapping:
                ma_dv_chuan, ten_dv_chuan, nhom_chinh = mapping
            else:
                ma_dv_chuan = ten_dv
                ten_dv_chuan = ten_dv
                nhom_chinh = "HCC"  # Mặc định thuộc HCC
                warnings.append(f"Dòng {idx+2}: Dịch vụ '{ten_dv}' chưa được định nghĩa trong danh mục. Tạm xếp vào HCC.")
                
            # Đọc các cột ngày bắt đầu và kết thúc (Từ cột 5 đến 10)
            tu_ngay = _safe_int(row.iloc[5]) if len(row) > 5 and pd.notna(row.iloc[5]) else 1
            tu_thang = _safe_int(row.iloc[6]) if len(row) > 6 and pd.notna(row.iloc[6]) else (int(thang[1:]) if thang.startswith('T') else 1)
            tu_nam = _safe_int(row.iloc[7]) if len(row) > 7 and pd.notna(row.iloc[7]) else nam_du_lieu
            
            den_ngay = _safe_int(row.iloc[8]) if len(row) > 8 and pd.notna(row.iloc[8]) else 30
            den_thang = _safe_int(row.iloc[9]) if len(row) > 9 and pd.notna(row.iloc[9]) else (int(thang[1:]) if thang.startswith('T') else 12)
            den_nam = _safe_int(row.iloc[10]) if len(row) > 10 and pd.notna(row.iloc[10]) else nam_du_lieu
            
            # Tính khoảng ngày và phân rã
            try:
                start_date = date(tu_nam, tu_thang, tu_ngay)
                end_date = date(den_nam, den_thang, den_ngay)
                if start_date > end_date:
                    start_date, end_date = end_date, start_date
            except Exception:
                # Fallback về cả tháng
                try:
                    thang_num = int(thang[1:]) if thang.startswith('T') else int(thang)
                    start_date = date(nam_du_lieu, thang_num, 1)
                    _, last_day = calendar.monthrange(nam_du_lieu, thang_num)
                    end_date = date(nam_du_lieu, thang_num, last_day)
                except Exception:
                    start_date = date(nam_du_lieu, 1, 1)
                    end_date = date(nam_du_lieu, 1, 31)
                    
            total_days = (end_date - start_date).days + 1
            if total_days <= 0:
                total_days = 1
                
            # Xác định bảng đích
            table_dest = f"transactions_{nhom_chinh.lower()}"
            if table_dest not in records_by_table:
                records_by_table[table_dest] = []
                
            # Phân rã thành từng ngày
            for day_idx in range(total_days):
                curr_date = start_date + timedelta(days=day_idx)
                curr_ngay = curr_date.day
                curr_thang = curr_date.month
                curr_nam = curr_date.year
                curr_thang_str = f"T{curr_thang:02d}"
                
                # Doanh thu chia đều số thực
                dt_day = dt / total_days
                # Sản lượng phân bổ làm tròn tích lũy
                cum_sl_current = round(sl * (day_idx + 1) / total_days)
                cum_sl_prev = round(sl * day_idx / total_days)
                sl_day = cum_sl_current - cum_sl_prev
                
                row_data = (
                    curr_thang_str, curr_nam, ma_bc, ma_dv_chuan, sl_day, dt_day, batch_id,
                    curr_ngay, curr_thang, curr_nam, curr_ngay, curr_thang, curr_nam,
                    _safe_int(stt_val)
                )
                records_by_table[table_dest].append(row_data)
                
        # 6. Chế độ ghi đè: Xóa dữ liệu cũ trùng bưu cục + dịch vụ + ngày phân rã cụ thể
        if mode == 'overwrite' and records_by_table:
            logger.info("Chế độ ghi đè: Đang thực hiện xóa dữ liệu cũ trùng bưu cục, dịch vụ và ngày phân rã cụ thể...")
            conn_del = _get_db_connection(db_path)
            try:
                cursor_del = conn_del.cursor()
                for target_tbl, rows in records_by_table.items():
                    # Xóa theo khóa: ten_dich_vu, ma_buu_cuc, tu_ngay, tu_thang, tu_nam
                    # (Lưu ý: r[3]=ten_dich_vu, r[2]=ma_buu_cuc, r[7]=tu_ngay, r[8]=tu_thang, r[9]=tu_nam)
                    delete_keys = set((r[3], r[2], r[7], r[8], r[9]) for r in rows)
                    cursor_del.executemany(f"""
                        DELETE FROM {target_tbl}
                        WHERE ten_dich_vu = ? AND ma_buu_cuc = ? AND tu_ngay = ? AND tu_thang = ? AND tu_nam = ?
                    """, list(delete_keys))
                conn_del.commit()
            except Exception as ex_del:
                logger.error(f"Lỗi khi xóa dữ liệu cũ trong chế độ ghi đè: {ex_del}", exc_info=True)
            finally:
                conn_del.close()
                
        # 7. Thực hiện ghi dữ liệu mới vào SQLite
        conn_ins = _get_db_connection(db_path)
        inserted = 0
        try:
            cursor_ins = conn_ins.cursor()
            for target_tbl, rows in records_by_table.items():
                insert_sql = f"""
                INSERT OR IGNORE INTO {target_tbl} 
                (thang_du_lieu, nam_du_lieu, ma_buu_cuc, ten_dich_vu, san_luong, doanh_thu, import_batch,
                 tu_ngay, tu_thang, tu_nam, den_ngay, den_thang, den_nam, stt)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """
                cursor_ins.executemany(insert_sql, rows)
                inserted += cursor_ins.rowcount
            conn_ins.commit()
        except Exception as ex_ins:
            logger.error(f"Lỗi khi insert dữ liệu dịch vụ mới: {ex_ins}", exc_info=True)
            raise
        finally:
            conn_ins.close()
            
        skipped = len([r for rows in records_by_table.values() for r in rows]) - inserted
        
        # 8. Tự động gộp số liệu cho các tháng bị ảnh hưởng
        if records_by_table:
            # Thu thập các tổ hợp (nhom_chinh, nam, thang) duy nhất để rebuild
            # nhom_chinh được suy ra từ tên bảng target_tbl (bỏ "transactions_")
            affected_months = set()
            for target_tbl, rows in records_by_table.items():
                nhom_ch = target_tbl.replace("transactions_", "").upper()
                for r in rows:
                    # r[9] = tu_nam, r[8] = tu_thang
                    affected_months.add((nhom_ch, r[9], r[8]))
                    
            for nhom_ch, nam_val, thang_val in affected_months:
                try:
                    logger.info(f"Tự động gộp số liệu cho {nhom_ch} - tháng T{thang_val:02d}/{nam_val}...")
                    _auto_aggregate_after_import(db_path, nhom_ch, [(nam_val, thang_val)])
                except Exception as ex_ref:
                    logger.error(f"Lỗi tự động refresh {nhom_ch} cho T{thang_val:02d}/{nam_val}: {ex_ref}")
                    
        return {
            'batch_id': batch_id,
            'file': filename,
            'thang': thang,
            'total_rows': total_rows,
            'inserted': inserted,
            'skipped': skipped,
            'warnings': warnings
        }
    except Exception as e:
        logger.error(f"Lỗi khi import {service_type}: {e}", exc_info=True)
        return {
            'batch_id': batch_id,
            'file': filename,
            'thang': thang,
            'total_rows': 0,
            'inserted': 0,
            'skipped': 0,
            'warnings': [str(e)]
        }


def import_plan_excel(db_path, excel_path, import_batch=None, mode='append'):
    """
    Import file Excel Kế hoạch vào bảng plans.
    Format mong đợi: Năm | Tháng | Nhóm DV | Tên DV | Mã BC | KH Doanh thu | KH Sản lượng
    """
    import pandas as pd
    filename = os.path.basename(excel_path)
    logger.info(f"Bắt đầu import Kế hoạch: {filename}")
    
    batch_id = import_batch or (datetime.now().strftime('%Y%m%d_%H%M%S') + '_PLAN_' + filename)
    
    try:
        engine = "openpyxl" if str(excel_path).endswith(".xlsx") else "xlrd"
        df = pd.read_excel(excel_path, engine=engine)
        
        if len(df.columns) < 5:
            return {'batch_id': batch_id, 'file': filename, 'thang': 'N/A', 'total_rows': 0, 'inserted': 0, 'skipped': 0, 'warnings': [f"File Kế hoạch không đủ số cột cơ bản. Số cột: {len(df.columns)}"]}
            

        # Nếu là overwrite, xóa kế hoạch cũ của năm/dịch vụ/đơn vị tương ứng
        if mode == 'overwrite' and len(df) > 0:
            logger.info("Chế độ ghi đè Kế hoạch: Đang xác định các năm, nhóm dịch vụ và bưu cục để xóa...")
            try:
                to_delete = set()
                valid_groups_local = {'BCCP', 'HCC', 'TCBC', 'PPBL', 'Hành chính công', 'Tài chính Bưu chính', 'Phân phối bán lẻ'}
                for idx, row in df.iterrows():
                    nam = _safe_int(row.iloc[0])
                    if nam == 0:
                        continue
                    nhom_chinh = clean_str_field(row.iloc[1])
                    nhom_dv = clean_str_field(row.iloc[2]) if len(df.columns) > 2 and pd.notna(row.iloc[2]) else None
                    ma_bc = clean_str_field(row.iloc[3]) if len(df.columns) > 3 else None
                    if nhom_chinh in valid_groups_local and ma_bc:
                        to_delete.add((nam, nhom_chinh, nhom_dv, ma_bc))
                
                if to_delete:
                    conn_del = _get_db_connection(db_path)
                    cursor_del = conn_del.cursor()
                    for nam, nhom_chinh, nhom_dv, ma_bc in to_delete:
                        if nhom_dv:
                            logger.info(f"Xóa kế hoạch cũ nam {nam} - {nhom_chinh}/{nhom_dv} bưu cục {ma_bc}...")
                            cursor_del.execute("""
                                DELETE FROM plans 
                                WHERE nam = ? AND nhom_chinh = ? AND nhom_dich_vu = ? AND ma_buu_cuc = ?
                            """, (nam, nhom_chinh, nhom_dv, ma_bc))
                        else:
                            logger.info(f"Xóa kế hoạch cũ nam {nam} - {nhom_chinh} bưu cục {ma_bc}...")
                            cursor_del.execute("""
                                DELETE FROM plans 
                                WHERE nam = ? AND nhom_chinh = ? AND nhom_dich_vu IS NULL AND ma_buu_cuc = ?
                            """, (nam, nhom_chinh, ma_bc))
                    conn_del.commit()
                    conn_del.close()
            except Exception as ex_del:
                logger.error(f"Lỗi khi xóa kế hoạch cũ: {ex_del}")

        conn = _get_db_connection(db_path)
        cursor = conn.cursor()
        
        insert_sql = """
        INSERT OR REPLACE INTO plans 
        (nam, thang, nhom_chinh, nhom_dich_vu, ma_buu_cuc, ke_hoach_doanh_thu, ke_hoach_san_luong)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """
        
        batch_buffer = []
        replaced = 0
        total_rows = 0
        warnings = []
        
        valid_groups = {'BCCP', 'HCC', 'TCBC', 'PPBL', 'Hành chính công', 'Tài chính Bưu chính', 'Phân phối bán lẻ'}
        
        # Tỷ lệ phân bổ 12 tháng (từ ty-le.xlsx)
        MONTHLY_RATIOS = [0.098266, 0.081857, 0.088313, 0.092610, 0.081116, 0.079334, 0.076506, 0.073452, 0.075493, 0.087763, 0.078089, 0.087201]
        
        last_nam = datetime.now().year
        for idx, row in df.iterrows():
            nam = _safe_int(row.iloc[0])
            if nam == 0:
                continue # Bỏ qua dòng trống
            last_nam = nam
                
            total_rows += 1
            nhom_chinh = clean_str_field(row.iloc[1])
            nhom_dv = clean_str_field(row.iloc[2]) if len(df.columns) > 2 and pd.notna(row.iloc[2]) else None
            ma_bc = clean_str_field(row.iloc[3]) if len(df.columns) > 3 else None
            kh_dt_nam = _safe_float(row.iloc[4]) if len(df.columns) > 4 else 0.0
            kh_sl = 0
            
            if nhom_chinh not in valid_groups:
                warnings.append(f"Dòng {idx+2}: Nhóm Chính không hợp lệ ('{nhom_chinh}')")
                continue
                
            for thang_idx, ratio in enumerate(MONTHLY_RATIOS, start=1):
                kh_dt_thang = kh_dt_nam * ratio
                batch_buffer.append((nam, thang_idx, nhom_chinh, nhom_dv, ma_bc, kh_dt_thang, kh_sl))
            
            if len(batch_buffer) >= BATCH_SIZE:
                cursor.executemany(insert_sql, batch_buffer)
                replaced += cursor.rowcount
                batch_buffer = []
                
        if batch_buffer:
            cursor.executemany(insert_sql, batch_buffer)
            replaced += cursor.rowcount
            
        skipped = total_rows - replaced
        
        # Ghi log
        cursor.execute("""
            INSERT INTO import_log (batch_id, file_name, thang_du_lieu, 
                                    so_dong_import, so_dong_trung, trang_thai, ghi_chu)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (batch_id, filename, 'PLAN', replaced, skipped, 'SUCCESS' if not warnings else 'SUCCESS_WITH_WARNINGS', '; '.join(warnings[:5]) if warnings else None))
        
        conn.commit()
        conn.close()

        # Tự động gộp số liệu cho tháng vừa nạp
        try:
            _auto_aggregate_after_import(db_path, 'PLAN', [(last_nam, 1)])
        except Exception as ex_ref:
            logger.error(f"Lỗi tự động refresh PLAN: {ex_ref}")

        return {
            'batch_id': batch_id,
            'file': filename,
            'thang': 'PLAN',
            'total_rows': total_rows,
            'inserted': replaced,
            'skipped': skipped,
            'warnings': warnings
        }
    except Exception as e:
        logger.error(f"Lỗi khi import Kế hoạch: {e}")
        return {'batch_id': batch_id, 'file': filename, 'thang': 'N/A', 'total_rows': 0, 'inserted': 0, 'skipped': 0, 'warnings': [str(e)]}

def import_phbc_excel(db_path, excel_path, import_batch=None, mode='append'):
    """
    Import file Excel dữ liệu Phát hành báo chí (PHBC).
    """
    import pandas as pd
    filename = os.path.basename(excel_path)
    logger.info(f"Bắt đầu import PHBC: {filename}")
    
    batch_id = import_batch or (datetime.now().strftime('%Y%m%d_%H%M%S') + '_PHBC_' + filename)
    
    try:
        engine = "openpyxl" if str(excel_path).endswith(".xlsx") else "xlrd"
        df = pd.read_excel(excel_path, engine=engine)
        
        # Xác định cột dựa vào tên
        cols = [str(c).lower().strip() for c in df.columns]
        
        bc_idx, dt_idx, thang_idx, nam_idx = None, None, None, None
        
        for i, c in enumerate(cols):
            if 'bưu cục' in c or 'buu_cuc' in c or 'buu cuc' in c or 'ma_bc' in c or 'mã bc' in c or 'bc' in c:
                if bc_idx is None: bc_idx = i
            if 'doanh thu' in c or 'doanh_thu' in c or 'dt' in c:
                if dt_idx is None: dt_idx = i
            if 'tháng' in c or 'thang' in c:
                if thang_idx is None: thang_idx = i
            if 'năm' in c or 'nam' in c or 'year' in c:
                if nam_idx is None: nam_idx = i
                
        if None in (bc_idx, dt_idx, thang_idx, nam_idx):
            warnings = [f"File PHBC thiếu cột. Các cột: {', '.join(df.columns)}"]
            return {'batch_id': batch_id, 'file': filename, 'thang': 'N/A', 'total_rows': 0, 'inserted': 0, 'skipped': 0, 'warnings': warnings}
            

        # Nếu là overwrite, xóa dữ liệu PHBC cũ của tháng/năm
        if mode == 'overwrite' and len(df) > 0:
            logger.info("Chế độ ghi đè PHBC: Đang xác định tháng/năm và bưu cục để xóa...")
            try:
                to_delete = set()
                for idx, row in df.iterrows():
                    bc_val = str(row.iloc[bc_idx]).strip().upper() if pd.notna(row.iloc[bc_idx]) else ""
                    thang_val = str(row.iloc[thang_idx]).strip() if pd.notna(row.iloc[thang_idx]) else ""
                    nam_val = _safe_int(row.iloc[nam_idx])
                    
                    if bc_val and thang_val:
                        if not thang_val.startswith("T"):
                            try:
                                thang_num = int(float(thang_val))
                                thang_val = f"T{thang_num:02d}"
                            except:
                                continue
                        if nam_val == 0:
                            nam_val = datetime.now().year
                        to_delete.add((thang_val, nam_val, bc_val))
                
                if to_delete:
                    conn_del = _get_db_connection(db_path)
                    cursor_del = conn_del.cursor()
                    for thang_val, nam_val, bc_val in to_delete:
                        logger.info(f"Xóa PHBC cũ của bưu cục {bc_val} tháng {thang_val}/{nam_val}...")
                        cursor_del.execute("""
                            DELETE FROM transactions_phbc
                            WHERE thang_du_lieu = ? AND nam_du_lieu = ? AND ma_buu_cuc = ?
                        """, (thang_val, nam_val, bc_val))
                    conn_del.commit()
                    conn_del.close()
            except Exception as ex_del:
                logger.error(f"Lỗi khi xóa PHBC cũ: {ex_del}")

        conn = _get_db_connection(db_path)
        cursor = conn.cursor()
        
        insert_sql = """
        INSERT OR REPLACE INTO transactions_phbc 
        (thang_du_lieu, nam_du_lieu, ma_buu_cuc, doanh_thu, import_batch)
        VALUES (?, ?, ?, ?, ?)
        """
        
        batch_buffer = []
        inserted = 0
        total_rows = 0
        warnings = []
        first_thang = "T01"
        last_nam = datetime.now().year
        
        for idx, row in df.iterrows():
            total_rows += 1
            bc_val = str(row.iloc[bc_idx]).strip().upper() if pd.notna(row.iloc[bc_idx]) else ""
            dt_val = _safe_float(row.iloc[dt_idx])
            thang_val = str(row.iloc[thang_idx]).strip() if pd.notna(row.iloc[thang_idx]) else ""
            nam_val = _safe_int(row.iloc[nam_idx])
            
            if not bc_val:
                continue
                
            if thang_val.startswith("T"):
                pass
            else:
                try:
                    thang_num = int(float(thang_val))
                    thang_val = f"T{thang_num:02d}"
                except ValueError:
                    warnings.append(f"Dòng {idx+2}: Tháng không hợp lệ ('{thang_val}')")
                    continue
                    
            if total_rows == 1:
                first_thang = thang_val
                
            if nam_val == 0:
                nam_val = datetime.now().year
                
            batch_buffer.append((thang_val, nam_val, bc_val, dt_val, batch_id))
            
            if len(batch_buffer) >= BATCH_SIZE:
                cursor.executemany(insert_sql, batch_buffer)
                inserted += cursor.rowcount
                batch_buffer = []
                
        if batch_buffer:
            cursor.executemany(insert_sql, batch_buffer)
            inserted += cursor.rowcount
            
        skipped = total_rows - inserted
        
        cursor.execute("""
            INSERT INTO import_log (batch_id, file_name, thang_du_lieu, 
                                    so_dong_import, so_dong_trung, trang_thai, ghi_chu)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (batch_id, filename, first_thang, inserted, skipped, 'SUCCESS' if not warnings else 'SUCCESS_WITH_WARNINGS', '; '.join(warnings[:5]) if warnings else None))
        
        conn.commit()
        conn.close()

        # Tự động gộp số liệu cho tháng vừa nạp
        try:
            thang_int = int(thang[1:]) if thang.startswith('T') else int(thang)
            _auto_aggregate_after_import(db_path, service_type, [(nam_du_lieu, thang_int)])
        except Exception as ex_ref:
            logger.error(f"Lỗi tự động refresh {service_type}: {ex_ref}")

        return {
            'batch_id': batch_id,
            'file': filename,
            'thang': first_thang,
            'total_rows': total_rows,
            'inserted': inserted,
            'skipped': skipped,
            'warnings': warnings
        }
    except Exception as e:
        logger.error(f"Lỗi khi import PHBC: {e}")
        return {'batch_id': batch_id, 'file': filename, 'thang': 'N/A', 'total_rows': 0, 'inserted': 0, 'skipped': 0, 'warnings': [str(e)]}
